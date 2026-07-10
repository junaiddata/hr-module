from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from .forms import EmployeeForm
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django_countries import countries
from django.http import JsonResponse
from django.utils import timezone

def role_required(allowed_roles=[]):
    def decorator(view_func):
        def wrapper_func(request, *args, **kwargs):
            if not request.user.is_authenticated:
                from django.contrib.auth.views import redirect_to_login
                return redirect_to_login(request.get_full_path())
            try:
                user_role = request.user.role.role
            except Exception:
                return render(request, 'hr_de/unauthorized.html', status=403)
            if user_role in allowed_roles:
                return view_func(request, *args, **kwargs)
            return render(request, 'hr_de/unauthorized.html', status=403)
        return wrapper_func
    return decorator


def _is_hr_department(dept):
    """The 'HR' department is special: its head gets full HR access, so we store
    their Role as 'HR' (scoped to the HR dept) instead of 'Head'."""
    return bool(dept) and (dept.name or '').strip().upper() == 'HR'


def _head_role_for(dept):
    """Role value to assign when making someone head of `dept`."""
    return 'HR' if _is_hr_department(dept) else 'Head'


def _head_department(request):
    """Department this user is head of (role='Head'), or None."""
    try:
        role = request.user.role
    except Exception:
        return None
    if role and role.role == 'Head' and role.department:
        return role.department
    return None


@login_required
def employee_list(request):
    query         = (request.GET.get('q') or '').strip()
    status_filter = request.GET.get('employee_status')
    active_filter = request.GET.get('is_active')
    dept_filter   = (request.GET.get('department') or '').strip()
    mol_filter    = (request.GET.get('mol') or '').strip()

    employees = Employee.objects.select_related('department', 'mol').all()

    # Restrict employees for department head
    is_head = hasattr(request.user, 'role') and request.user.role.role == 'Head'
    if is_head:
        dept = request.user.role.department
        employees = employees.filter(department=dept) if dept else employees.none()

    # Broad text search across the common identifying fields
    if query:
        employees = employees.filter(
            Q(emp_name__icontains=query) |
            Q(emp_id__icontains=query) |
            Q(designation__icontains=query) |
            Q(contact_number__icontains=query) |
            Q(job_location__icontains=query) |
            Q(nationality__icontains=query) |
            Q(passport_number__icontains=query) |
            Q(eid_number__icontains=query) |
            Q(department__name__icontains=query) |
            Q(mol__mol__icontains=query)
        )

    if dept_filter:
        employees = employees.filter(department_id=dept_filter)

    if mol_filter:
        employees = employees.filter(mol_id=mol_filter)

    if status_filter:
        employees = employees.filter(employee_status=status_filter)

    # Active/inactive filter
    if active_filter:
        employees = employees.filter(is_active=(active_filter == 'true'))

    employees = employees.order_by('emp_name')

    # Pass status choices to template
    status_choices = Employee.EMPLOYEE_STATUS_CHOICES
    active_count = employees.filter(is_active=True).count()
    inactive_count = employees.filter(is_active=False).count()

    # Dropdown sources (heads only see their own department)
    if is_head and request.user.role.department:
        departments = Department.objects.filter(pk=request.user.role.department_id)
    else:
        departments = Department.objects.all().order_by('name')
    mols = Mol.objects.all().order_by('mol')

    # Role-based template rendering for employee list
    template_name = 'hr_de/employee_list_head.html' if is_head else 'hr_de/employee_list.html'

    return render(request, template_name, {
        'employees': employees,
        'status_choices': status_choices,
        'query': query,
        'selected_status': status_filter,
        'selected_active': active_filter,
        'selected_dept': dept_filter,
        'selected_mol': mol_filter,
        'departments': departments,
        'mols': mols,
        'active_count': active_count,
        'inactive_count': inactive_count,
    })

@login_required
def employee_create(request):
    # Inline role check — Head users cannot create employees
    try:
        user_role = request.user.role.role
    except Exception:
        user_role = None

    if user_role not in ['HR']:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        data = request.POST
        files = request.FILES

        # Helper function to handle empty dates
        def parse_date(date_str):
            return date_str if date_str else None


        # Get the Mol instance if mol ID was provided
        mol_id = data.get('mol')
        mol_instance = None
        is_active = data.get('is_active') == 'on'
        if mol_id:
            try:
                mol_instance = Mol.objects.get(pk=mol_id)
            except Mol.DoesNotExist:
                pass

        dept_obj = Department.objects.filter(pk=data.get('department')).first()
        employee = Employee.objects.create(
            emp_name=data.get('emp_name'),
            emp_id=data.get('emp_id'),
            photo=files.get('photo'),
            dob=parse_date(data.get('dob')),
            gender=data.get('gender', 'M'),
            nationality=data.get('nationality', 'Indian'),
            contact_number=data.get('contact_number'),
            department=dept_obj,
            designation=data.get('designation'),
            mol=mol_instance,
            joining_date=parse_date(data.get('joining_date')),
            emp_salary= data.get('emp_salary') or None,
            job_location=data.get('job_location'),
            employee_status=data.get('employee_status', 'Active'),
            is_active=is_active,

            passport=files.get('passport'),
            visa=files.get('visa'),
            labour_card=files.get('labour_card'),
            eid=files.get('eid'),
            insurance=files.get('insurance'),
            driving_license=files.get('driving_license'),

            passport_status=data.get('passport_status', 'With company'),
            passport_number=data.get('passport_number'),
            visa_number=data.get('visa_number'),
            eid_number=data.get('eid_number'),
            labour_card_number=data.get('labour_card_number'),
            labour_number=data.get('labour_number'),
            insurance_number=data.get('insurance_number'),
            driving_license_number=data.get('driving_license_number'),

            bank_name=data.get('bank_name'),
            iban=data.get('iban'),
            routing_code=data.get('routing_code'),

            visa_expiry=parse_date(data.get('visa_expiry')),
            passport_expiry=parse_date(data.get('passport_expiry')),
            labour_card_expiry=parse_date(data.get('labour_card_expiry')),
            eid_expiry=parse_date(data.get('eid_expiry')),
            insurance_expiry=parse_date(data.get('insurance_expiry')),
            driving_license_expiry=parse_date(data.get('driving_license_expiry')),
        )

        # Save salary structure if any component is provided
        def _float(key):
            try:
                return float(data.get(key, 0) or 0)
            except ValueError:
                return 0.0

        basic     = _float('basic')
        hra       = _float('hra')
        transport = _float('transport')
        fuel      = _float('fuel')
        others    = _float('others')
        if any([basic, hra, transport, fuel, others]):
            SalaryStructure.objects.create(
                employee=employee,
                basic=basic,
                hra=hra,
                transport=transport,
                fuel=fuel,
                others=others,
                updated_by=request.user,
            )

        # Create login account if username provided
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        if username and password:
            if User.objects.filter(username=username).exists():
                messages.error(request, f"Username '{username}' is already taken. Employee created without login account.")
            else:
                new_user = User.objects.create_user(username=username, password=password, first_name=data.get('emp_name', ''))
                employee.user = new_user
                employee.save()

                # Make department head if checked (head of the HR dept becomes full HR)
                if data.get('make_head') == 'on' and dept_obj:
                    Role.objects.update_or_create(
                        user=new_user,
                        defaults={'role': _head_role_for(dept_obj), 'department': dept_obj}
                    )

        return redirect('employee_list')

    return render(request, 'hr_de/employee_form_raw.html', {
        'mols': Mol.objects.all(),
        'countries': countries,
        'departments': Department.objects.all().order_by('name'),
    })


@login_required
def employee_edit(request, pk):
    try:
        _edit_role = request.user.role.role
    except Exception:
        _edit_role = None
    if _edit_role not in ['HR']:
        return render(request, 'hr_de/unauthorized.html', status=403)

    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        data = request.POST
        files = request.FILES

        is_active = data.get('is_active') == 'on'

        # Helper function to handle empty dates
        def parse_date(date_str):
            return date_str if date_str else None

        # Get the Mol instance if mol ID was provided
        mol_id = data.get('mol')
        mol_instance = None
        if mol_id:
            try:
                mol_instance = Mol.objects.get(pk=mol_id)
            except Mol.DoesNotExist:
                pass

        employee.emp_name = data.get('emp_name')
        employee.emp_id = data.get('emp_id')
        employee.dob = parse_date(data.get('dob'))
        employee.gender = data.get('gender')
        employee.nationality = data.get('nationality')
        employee.contact_number = data.get('contact_number')
        dept_obj = Department.objects.filter(pk=data.get('department')).first()
        employee.department = dept_obj
        employee.designation = data.get('designation')
        employee.mol = mol_instance
        employee.joining_date = parse_date(data.get('joining_date'))
        employee.emp_salary = data.get('emp_salary') or None
        employee.job_location = data.get('job_location')
        employee.employee_status = data.get('employee_status')
        employee.is_active = is_active

        employee.passport_status = data.get('passport_status')
        employee.passport_number = data.get('passport_number')
        employee.visa_number = data.get('visa_number')
        employee.eid_number = data.get('eid_number')
        employee.labour_card_number = data.get('labour_card_number')
        employee.labour_number = data.get('labour_number')
        employee.insurance_number = data.get('insurance_number')
        employee.driving_license_number = data.get('driving_license_number')

        employee.bank_name = data.get('bank_name')
        employee.iban = data.get('iban')
        employee.routing_code = data.get('routing_code')

        employee.visa_expiry = parse_date(data.get('visa_expiry'))
        employee.passport_expiry = parse_date(data.get('passport_expiry'))
        employee.labour_card_expiry = parse_date(data.get('labour_card_expiry'))
        employee.eid_expiry = parse_date(data.get('eid_expiry'))
        employee.insurance_expiry = parse_date(data.get('insurance_expiry'))
        employee.driving_license_expiry = parse_date(data.get('driving_license_expiry'))

        # Handle file uploads
        if files.get('photo'):
            employee.photo = files.get('photo')
        if files.get('passport'):
            employee.passport = files.get('passport')
        if files.get('visa'):
            employee.visa = files.get('visa')
        if files.get('labour_card'):
            employee.labour_card = files.get('labour_card')
        if files.get('eid'):
            employee.eid = files.get('eid')
        if files.get('insurance'):
            employee.insurance = files.get('insurance')
        if files.get('driving_license'):
            employee.driving_license = files.get('driving_license')

        # Create/update login account
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        make_head = data.get('make_head') == 'on'

        if username:
            if employee.user:
                # Update existing user
                emp_user = employee.user
                if emp_user.username != username:
                    if not User.objects.filter(username=username).exclude(pk=emp_user.pk).exists():
                        emp_user.username = username
                if password:
                    emp_user.set_password(password)
                emp_user.save()
            else:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' already taken.")
                else:
                    emp_user = User.objects.create_user(username=username, password=password or 'changeme123', first_name=employee.emp_name)
                    employee.user = emp_user

        employee.save()

        # Update salary structure if any component supplied
        def _float_edit(key):
            try:
                return float(data.get(key, 0) or 0)
            except ValueError:
                return 0.0

        basic     = _float_edit('basic')
        hra       = _float_edit('hra')
        transport = _float_edit('transport')
        fuel      = _float_edit('fuel')
        others    = _float_edit('others')
        if any([basic, hra, transport, fuel, others]):
            SalaryStructure.objects.update_or_create(
                employee=employee,
                defaults={
                    'basic': basic, 'hra': hra,
                    'transport': transport, 'fuel': fuel, 'others': others,
                    'updated_by': request.user,
                }
            )

        # Handle head promotion/demotion (head of the HR dept becomes full HR)
        if employee.user:
            emp_user = employee.user
            if make_head and dept_obj:
                Role.objects.update_or_create(
                    user=emp_user,
                    defaults={'role': _head_role_for(dept_obj), 'department': dept_obj}
                )
            elif not make_head:
                # Remove the department-head role if unchecked. Only touches roles
                # scoped to a department, so company-level HR (dept=None) is safe.
                Role.objects.filter(
                    user=emp_user, role__in=['Head', 'HR'], department__isnull=False
                ).delete()

        return redirect('employee_list')

    # Pre-check the "Make Department Head" box for anyone who heads a department
    # (regular heads use role='Head'; the HR-dept head uses role='HR', dept set).
    is_head = False
    if employee.user:
        is_head = Role.objects.filter(
            user=employee.user, role__in=['Head', 'HR'], department__isnull=False
        ).exists()

    return render(request, 'hr_de/employee_form_raw.html', {
        'employee': employee,
        'mols': Mol.objects.all(),
        'countries': countries,
        'departments': Department.objects.all().order_by('name'),
        'is_head': is_head,
    })



@login_required
def employee_detail(request, pk):
    from datetime import date

    employee = get_object_or_404(Employee, pk=pk)

    try:
        viewer_role = request.user.role.role
    except Exception:
        viewer_role = None
    is_own = bool(employee.user_id) and employee.user_id == request.user.id

    # Self-service employees (no HR/MD/Head role) may only view their own profile
    if viewer_role not in ('HR', 'MD', 'Head') and not is_own:
        return render(request, 'hr_de/unauthorized.html', status=403)

    today = date.today()

    def _years_since(d):
        """Whole years between date d and today (age / tenure)."""
        if not d:
            return None
        return today.year - d.year - ((today.month, today.day) < (d.month, d.day))

    def _expiry(d):
        """Return (status, days_left) for a document expiry date.
        status: expired / critical (<=30d) / warning (<=90d) / valid / none."""
        if not d:
            return ('none', None)
        days = (d - today).days
        if days < 0:
            return ('expired', days)
        if days <= 30:
            return ('critical', days)
        if days <= 90:
            return ('warning', days)
        return ('valid', days)

    # Unified document list: label, id number, expiry date, uploaded file, status
    doc_specs = [
        ('Passport',        employee.passport_number,        employee.passport_expiry,        employee.passport),
        ('Visa',            employee.visa_number,            employee.visa_expiry,            employee.visa),
        ('Labour Card',     employee.labour_card_number,     employee.labour_card_expiry,     employee.labour_card),
        ('Emirates ID',     employee.eid_number,             employee.eid_expiry,             employee.eid),
        ('Insurance',       employee.insurance_number,       employee.insurance_expiry,       employee.insurance),
        ('Driving License', employee.driving_license_number, employee.driving_license_expiry, employee.driving_license),
    ]
    documents = []
    expiry_alerts = 0
    for label, number, expiry, file in doc_specs:
        status, days = _expiry(expiry)
        if status in ('expired', 'critical', 'warning'):
            expiry_alerts += 1
        documents.append({
            'label': label,
            'number': number,
            'expiry': expiry,
            'file': file,
            'status': status,
            'days': days,
            'abs_days': abs(days) if days is not None else None,
        })

    return render(request, 'hr_de/employee_detail.html', {
        'employee': employee,
        'documents': documents,
        'expiry_alerts': expiry_alerts,
        'age': _years_since(employee.dob),
        'tenure': _years_since(employee.joining_date),
        'is_own': is_own,
        # Related "Other Records" — HR/MD only
        'other_records': (
            employee.other_records.select_related('uploaded_by').all()
            if _hr_or_md(request) else []
        ),
        # Vehicles assigned to this employee (fleet register)
        'related_vehicles': employee.vehicles.all().order_by('name'),
        'can_manage_vehicles': _hr_or_md(request),
        # Company properties assigned to this employee
        'related_properties': employee.properties.all().order_by('name'),
    })


@login_required
@require_POST
def upload_my_photo(request):
    """Employee updates ONLY their own profile photo — no other fields."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    photo = request.FILES.get('photo')
    if not photo:
        messages.error(request, 'Please choose an image file to upload.')
        return redirect('employee_detail', pk=employee.pk)

    content_type = getattr(photo, 'content_type', '') or ''
    if not content_type.startswith('image/'):
        messages.error(request, 'Only image files (JPG, PNG, WEBP) are allowed.')
        return redirect('employee_detail', pk=employee.pk)

    employee.photo = photo
    employee.save(update_fields=['photo', 'updated_at'])
    messages.success(request, 'Your profile photo has been updated.')
    return redirect('employee_detail', pk=employee.pk)


@login_required
def my_profile(request):
    """Send any logged-in user to their own profile page."""
    try:
        employee = request.user.employee
    except Exception:
        messages.error(request, 'Your account is not linked to an employee record.')
        return redirect('home')
    return redirect('employee_detail', pk=employee.pk)


@login_required
def birthdays(request):
    """HR view of employee birthdays — defaults to this + next month, or a single
    month picked via ?month=1-12. Supports ?export=pdf. Active employees only."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    from datetime import date
    import calendar
    from .whatsapp import whatsapp_configured

    today = date.today()

    # Employees already wished this year (to show "Sent" state on today's cards)
    wished_ids = set(
        BirthdayWish.objects.filter(year=today.year, status='sent')
        .values_list('employee_id', flat=True)
    )

    def _occurrence(dob, year):
        try:
            return date(year, dob.month, dob.day)
        except ValueError:                   # Feb 29 in a non-leap year
            return date(year, dob.month, 28)

    def _row(e, bday):
        return {
            'employee':   e,
            'day':        e.dob.day,
            'bday':       bday,
            'turning':    bday.year - e.dob.year,
            'days_until': (bday - today).days,
            'is_today':   e.dob.month == today.month and e.dob.day == today.day,
            'wished':     e.pk in wished_ids,
        }

    def build(month, year):
        qs = Employee.objects.filter(is_active=True, dob__month=month).select_related('department')
        rows = [_row(e, _occurrence(e.dob, year)) for e in qs]
        rows.sort(key=lambda r: r['day'])
        return rows

    # ── Single month picker (?month=1-12) ───────────────────────────
    try:
        sel_month = int(request.GET.get('month') or 0)
    except (ValueError, TypeError):
        sel_month = 0
    month_mode = 1 <= sel_month <= 12

    if month_mode:
        # Use the upcoming occurrence of that month for "turns X"
        year = today.year if sel_month >= today.month else today.year + 1
        rows = build(sel_month, year)
        sections = [(f"{calendar.month_name[sel_month]} {year}", rows)]
        total = len(rows)
    else:
        # ── Default: this month + next month ────────────────────────
        this_month = today.month
        next_month = 1 if this_month == 12 else this_month + 1
        this_year  = today.year
        next_year  = this_year + 1 if next_month == 1 else this_year
        this_rows = build(this_month, this_year)
        next_rows = build(next_month, next_year)
        sections = [
            (f"{calendar.month_name[this_month]} (This Month)", this_rows),
            (f"{calendar.month_name[next_month]} (Next Month)", next_rows),
        ]
        total = len(this_rows) + len(next_rows)

    if request.GET.get('export') == 'pdf':
        subtitle = sections[0][0] if month_mode else 'This month and next month'
        return _birthdays_pdf(sections, today, subtitle)

    # Month dropdown options
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]

    return render(request, 'hr_de/birthdays.html', {
        'sections':      sections,
        'total':         total,
        'today':         today,
        'month_mode':    month_mode,
        'selected_month': sel_month if month_mode else 0,
        'months':        months,
        'whatsapp_ready': whatsapp_configured(),
    })


@login_required
def birthday_wish_send(request):
    """AJAX: HR sends a WhatsApp birthday wish to one employee."""
    if not _hr_only(request):
        return JsonResponse({'ok': False, 'error': 'Not authorized.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    from .whatsapp import send_birthday_wish, whatsapp_configured
    import json as _json
    try:
        data = _json.loads(request.body or '{}')
    except ValueError:
        data = {}
    emp_pk = data.get('employee_pk') or request.POST.get('employee_pk')

    emp = Employee.objects.filter(pk=emp_pk, is_active=True).first()
    if not emp:
        return JsonResponse({'ok': False, 'error': 'Employee not found.'}, status=404)
    if not whatsapp_configured():
        return JsonResponse({
            'ok': False,
            'error': 'WhatsApp is not configured yet — add the token and phone number id in settings.',
        }, status=400)

    ok, info = send_birthday_wish(emp, sent_by=request.user)
    if ok:
        return JsonResponse({'ok': True, 'message': f'Birthday wish sent to {emp.emp_name}.'})
    return JsonResponse({'ok': False, 'error': info}, status=400)


@login_required
def birthday_wish_log(request):
    """HR view of the WhatsApp birthday-wish send history (sent + failed)."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    from .whatsapp import whatsapp_configured

    base = (BirthdayWish.objects
            .select_related('employee', 'employee__department', 'sent_by'))

    # Counts across everything (for the filter tab badges)
    sent_count = base.filter(status='sent').count()
    failed_count = base.filter(status='failed').count()

    status = request.GET.get('status')
    wishes = base.filter(status=status) if status in ('sent', 'failed') else base

    return render(request, 'hr_de/birthday_wish_log.html', {
        'wishes':        list(wishes),
        'sent_count':    sent_count,
        'failed_count':  failed_count,
        'total':         sent_count + failed_count,
        'status':        status or '',
        'whatsapp_ready': whatsapp_configured(),
    })


def _birthdays_pdf(sections, today, subtitle=''):
    """Render birthday `sections` as a polished, card-style PDF with employee
    photos — mirroring the website design (ReportLab + Pillow)."""
    import io, os
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from PIL import Image as PILImage, ImageDraw, ImageFont

    BRAND = colors.HexColor('#2208cd')
    PINK  = colors.HexColor('#ec4899')
    INK   = colors.HexColor('#0f172a')
    MUTE  = colors.HexColor('#64748b')

    def _avatar(employee, bg=(255, 255, 255), size=120):
        """Circular avatar PNG (photo centre-cropped, or a coloured initial),
        composited onto the row background `bg` so corners stay clean on any
        ReportLab version (no reliance on PNG alpha)."""
        base  = PILImage.new('RGBA', (size, size), (0, 0, 0, 0))
        drawn = False
        if getattr(employee, 'photo', None):
            try:
                path = employee.photo.path
                if os.path.exists(path):
                    ph = PILImage.open(path).convert('RGB')
                    w, h = ph.size
                    m = min(w, h)
                    ph = ph.crop(((w - m) // 2, (h - m) // 2, (w - m) // 2 + m, (h - m) // 2 + m)).resize((size, size))
                    base.paste(ph, (0, 0))
                    drawn = True
            except Exception:
                drawn = False
        if not drawn:
            d = ImageDraw.Draw(base)
            d.ellipse((0, 0, size - 1, size - 1), fill=(34, 8, 205, 255))
            letter = ((employee.emp_name or '?').strip()[:1] or '?').upper()
            font = None
            for fname in ('arialbd.ttf', 'arial.ttf', 'DejaVuSans-Bold.ttf'):
                try:
                    font = ImageFont.truetype(fname, int(size * 0.46)); break
                except Exception:
                    continue
            if font is None:
                font = ImageFont.load_default()
            bb = d.textbbox((0, 0), letter, font=font)
            d.text(((size - (bb[2] - bb[0])) / 2 - bb[0], (size - (bb[3] - bb[1])) / 2 - bb[1]),
                   letter, font=font, fill=(255, 255, 255, 255))
        mask = PILImage.new('L', (size, size), 0)
        ImageDraw.Draw(mask).ellipse((0, 0, size - 1, size - 1), fill=255)
        # Composite the masked avatar onto the row background → flat RGB, clean edges
        canvas = PILImage.new('RGB', (size, size), bg)
        canvas.paste(base, (0, 0), mask)
        buf = BytesIO(); canvas.save(buf, 'PNG'); buf.seek(0)
        return buf

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        title='Upcoming Birthdays',
    )
    styles = getSampleStyleSheet()
    name_style = ParagraphStyle('name', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=INK, leading=12)
    role_style = ParagraphStyle('role', parent=styles['Normal'], fontSize=8, textColor=MUTE, leading=10)
    date_style = ParagraphStyle('date', parent=styles['Normal'], fontSize=10, textColor=INK, alignment=2, leading=12)
    turn_style = ParagraphStyle('turn', parent=styles['Normal'], fontSize=8, textColor=MUTE, alignment=2, leading=11)
    hdr_style  = ParagraphStyle('hdr', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, textColor=colors.white, leading=14)

    # Title band
    elems = []
    band = Table([[Paragraph('Upcoming Birthdays', ParagraphStyle('tt', fontName='Helvetica-Bold', fontSize=17, textColor=colors.white))]],
                 colWidths=[doc.width])
    band.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BRAND),
        ('TOPPADDING', (0, 0), (-1, -1), 12), ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
    ]))
    elems.append(band)
    elems.append(Spacer(1, 5))
    elems.append(Paragraph((subtitle + ' &nbsp;&middot;&nbsp; ' if subtitle else '') + 'Generated on ' + today.strftime('%d %b %Y'),
                           ParagraphStyle('sub', fontSize=9, textColor=MUTE)))
    elems.append(Spacer(1, 12))

    def status_para(r):
        if r['is_today']:
            txt, col = 'Today', PINK
        elif r['days_until'] < 0:
            txt, col = 'Celebrated', MUTE
        elif r['days_until'] <= 7:
            txt, col = ('In %d day%s' % (r['days_until'], '' if r['days_until'] == 1 else 's'), colors.HexColor('#b45309'))
        else:
            txt, col = ('In %d days' % r['days_until'], colors.HexColor('#1d4ed8'))
        return Paragraph('<b>%s</b>' % txt, ParagraphStyle('st', fontSize=8, textColor=col, alignment=2, leading=11))

    def section(title, rows):
        sh = Table([[Paragraph('%s  (%d)' % (title, len(rows)), hdr_style)]], colWidths=[doc.width])
        sh.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), PINK),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elems.append(sh)
        elems.append(Spacer(1, 6))
        if not rows:
            elems.append(Paragraph('No birthdays.', role_style))
            elems.append(Spacer(1, 12))
            return
        data = []
        for i, r in enumerate(rows):
            e = r['employee']
            row_bg = (255, 255, 255) if i % 2 == 0 else (253, 242, 248)  # matches ROWBACKGROUNDS
            meta = e.designation or '—'
            if e.department:
                meta += ' &middot; %s' % e.department.name
            name_cell = [Paragraph(e.emp_name, name_style), Paragraph(meta, role_style)]
            date_cell = [
                Paragraph('<font color="#ec4899"><b>%s</b></font> %s' % (r['bday'].strftime('%d'), r['bday'].strftime('%b')), date_style),
                Paragraph('Turns %d' % r['turning'], turn_style),
            ]
            data.append([Image(_avatar(e, bg=row_bg), width=1.0 * cm, height=1.0 * cm), name_cell, date_cell, status_para(r)])
        tbl = Table(data, colWidths=[1.3 * cm, doc.width - 1.3 * cm - 2.6 * cm - 3.0 * cm, 2.6 * cm, 3.0 * cm])
        tbl.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#fdf2f8')]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.4, colors.HexColor('#f6d9e8')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elems.append(tbl)
        elems.append(Spacer(1, 14))

    for title, rows in sections:
        section(title, rows)

    doc.build(elems)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="upcoming_birthdays.pdf"'
    return resp


@login_required
def mol_list(request):
    mols = Mol.objects.all().order_by('mol')
    return render(request, 'hr_de/mols/mol_list.html', {'mols': mols})

@login_required
def mol_add(request):
    if request.method == 'POST':
        def _f(key): return request.POST.get(key, '').strip() or None
        mol = Mol(
            mol=request.POST.get('mol', '').strip(),
            established_year=_f('established_year'),
            company_code=_f('company_code'),
            wps_number=_f('wps_number'),
            iban=_f('iban'),
            trade_license_number=_f('trade_license_number'),
            trade_license_expiry=_f('trade_license_expiry'),
            tenancy_contract_expiry=_f('tenancy_contract_expiry'),
            establishment_card_expiry=_f('establishment_card_expiry'),
        )
        if 'license_document' in request.FILES:
            mol.license_document = request.FILES['license_document']
        if 'tenancy_contract' in request.FILES:
            mol.tenancy_contract = request.FILES['tenancy_contract']
        if 'establishment_card' in request.FILES:
            mol.establishment_card = request.FILES['establishment_card']
        mol.save()
        messages.success(request, f"MOL '{mol.mol}' created successfully.")
        return redirect('mol_list')
    return render(request, 'hr_de/mols/mol_form.html')

@login_required
def mol_edit(request, pk):
    mol = get_object_or_404(Mol, pk=pk)
    if request.method == 'POST':
        def _f(key): return request.POST.get(key, '').strip() or None
        mol.mol                       = request.POST.get('mol', '').strip()
        mol.established_year          = _f('established_year')
        mol.company_code             = _f('company_code')
        mol.wps_number               = _f('wps_number')
        mol.iban                     = _f('iban')
        mol.trade_license_number      = _f('trade_license_number')
        mol.trade_license_expiry      = _f('trade_license_expiry')
        mol.tenancy_contract_expiry   = _f('tenancy_contract_expiry')
        mol.establishment_card_expiry = _f('establishment_card_expiry')
        if 'license_document' in request.FILES:
            mol.license_document = request.FILES['license_document']
        if 'tenancy_contract' in request.FILES:
            mol.tenancy_contract = request.FILES['tenancy_contract']
        if 'establishment_card' in request.FILES:
            mol.establishment_card = request.FILES['establishment_card']
        mol.save()
        messages.success(request, f"MOL '{mol.mol}' updated successfully.")
        return redirect('mol_list')
    return render(request, 'hr_de/mols/mol_form.html', {'mol': mol})



@require_POST
def mol_delete(request, pk):
    mol = get_object_or_404(Mol, pk=pk)
    mol.delete()
    return redirect('mol_list')


# ── Department CRUD ──────────────────────────────────────────────────────────

@login_required
def department_list(request):
    all_depts = Department.objects.all().order_by('name')
    # Split out shops/stores into their own section (name starts with SHOP or STORE)
    shops, departments = [], []
    for d in all_depts:
        name = (d.name or '').strip().upper()
        (shops if name.startswith('SHOP') or name.startswith('STORE') else departments).append(d)
    return render(request, 'hr_de/departments/department_list.html', {
        'departments': departments,
        'shops': shops,
    })

@login_required
def department_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_support_staff = request.POST.get('is_support_staff') == 'on'
        if name:
            Department.objects.get_or_create(name=name, defaults={'is_support_staff': is_support_staff})
        return redirect('department_list')
    return render(request, 'hr_de/departments/department_form.html')

@login_required
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_support_staff = request.POST.get('is_support_staff') == 'on'
        if name:
            department.name = name
            department.is_support_staff = is_support_staff
            department.save()
        return redirect('department_list')
    return render(request, 'hr_de/departments/department_form.html', {'department': department})

@require_POST
@login_required
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    department.delete()
    return redirect('department_list')


#############################       LEAVE MANAGEMENT       #############################
def leave_type_list(request):
    leave_types = LeaveType.objects.all()
    return render(request, 'hr_de/leave_type/leave_type_list.html', {'leave_types': leave_types})

def leave_type_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        LeaveType.objects.create(name=name)
        return redirect('leave_type_list')
    return render(request, 'hr_de/leave_type/leave_type_form.html')

def leave_type_edit(request, pk):
    leave_type = get_object_or_404(LeaveType, pk=pk)

    if request.method == 'POST':
        leave_type.name = request.POST.get('name')
        leave_type.save()
        return redirect('leave_type_list')

    return render(request, 'hr_de/leave_type/leave_type_form.html', {'leave_type': leave_type})


def leave_type_delete(request, pk):
    leave_type = get_object_or_404(LeaveType, pk=pk)
    leave_type.delete()
    return redirect('leave_type_list')

@login_required
def add_leave(request):
    employees = Employee.objects.filter(is_active=True)

    # Restrict employees for department head
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        dept = request.user.role.department
        employees = employees.filter(department=dept) if dept else employees.none()

    leave_types = LeaveType.objects.all()

    if request.method == "POST":
        employee_id = request.POST.get("employee")
        leave_type_id = request.POST.get("leave_type")
        expected_from = request.POST.get("expected_from")
        expected_to = request.POST.get("expected_to")
        reported_to = request.POST.get("reported_to")
        reason = request.POST.get("reason")
        leave_application = request.FILES.get("leave_application")  # Get uploaded file

        Leave.objects.create(
            employee_id=employee_id,
            leave_type_id=leave_type_id,
            expected_from=expected_from,
            expected_to=expected_to,
            reported_to=reported_to,
            reason=reason,
            leave_application=leave_application,  # Save the file
            status='Pending'
        )
        return redirect('leave_history')

    # Role-based template rendering
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        template_name = 'hr_de/add_leave_head.html'
    else:
        template_name = 'hr_de/add_leave.html'

    return render(request, template_name, {
        'employees': employees,
        'leave_types': leave_types
    })


@login_required
def approve_leave(request, leave_id):
    leave = get_object_or_404(Leave, id=leave_id)
    user_role = getattr(request.user, 'role', None)

    if request.method == "POST":
        action = request.POST.get("action")
        rejection_reason = request.POST.get("rejection_reason", "").strip()

        if user_role and user_role.role == 'Head':
            if leave.status != 'Pending':
                messages.error(request, "This leave is no longer pending for Head approval.")
                return redirect('pending_leaves')
            if leave.employee.department != user_role.department:
                return render(request, 'hr_de/unauthorized.html', status=403)
            if action == 'approve':
                leave.status = 'Head_Approved'
                leave.head_approved_by = request.user
                leave.head_approved_at = timezone.now()
            elif action == 'reject':
                leave.status = 'Rejected'
                leave.head_approved_by = request.user
                leave.rejection_reason = rejection_reason

        elif user_role and user_role.role == 'MD':
            if leave.status != 'Head_Approved':
                messages.error(request, "This leave must be approved by the Head first.")
                return redirect('pending_leaves')
            if action == 'approve':
                leave.status = 'MD_Approved'
                leave.md_approved_by = request.user
                leave.md_approved_at = timezone.now()
            elif action == 'reject':
                leave.status = 'Rejected'
                leave.md_approved_by = request.user
                leave.rejection_reason = rejection_reason

        elif user_role and user_role.role == 'HR':
            if leave.status != 'MD_Approved':
                messages.error(request, "This leave must be approved by the MD first.")
                return redirect('pending_leaves')
            if action == 'approve':
                leave.status = 'Approved'
                leave.approved_by = request.user
            elif action == 'reject':
                leave.status = 'Rejected'
                leave.approved_by = request.user
                leave.rejection_reason = rejection_reason

        else:
            return render(request, 'hr_de/unauthorized.html', status=403)

        leave.save()
        return redirect('pending_leaves')

    return render(request, 'hr_de/approve_leave.html', {
        'leave': leave,
        'approver_role': user_role,
    })


@login_required
def pending_leaves(request):
    user_role = getattr(request.user, 'role', None)

    if user_role and user_role.role == 'Head':
        dept = user_role.department
        leaves = Leave.objects.filter(
            status='Pending',
            employee__department=dept
        ).select_related('employee', 'leave_type') if dept else Leave.objects.none()
    elif user_role and user_role.role == 'MD':
        leaves = Leave.objects.filter(
            status='Head_Approved'
        ).select_related('employee', 'leave_type', 'head_approved_by')
    elif user_role and user_role.role == 'HR':
        leaves = Leave.objects.filter(
            status='MD_Approved'
        ).select_related('employee', 'leave_type', 'head_approved_by', 'md_approved_by')
    else:
        leaves = Leave.objects.none()

    return render(request, 'hr_de/pending_leaves.html', {
        'leaves': leaves,
        'user_role': user_role,
    })

from django.db.models import Q

@login_required
def leave_history(request):
    query = request.GET.get('q') or ''
    leave_type_id = request.GET.get('leave_type')
    status_filter = request.GET.get('status')

    leaves = Leave.objects.select_related('employee', 'leave_type')

    # Filter based on role
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        dept = request.user.role.department
        leaves = leaves.filter(employee__department=dept) if dept else leaves.none()

    # Apply search filters
    if query:
        leaves = leaves.filter(
            Q(employee__emp_name__icontains=query) |
            Q(employee__emp_id__icontains=query)
        )

    if leave_type_id:
        leaves = leaves.filter(leave_type_id=leave_type_id)
        
    if status_filter:
        leaves = leaves.filter(status=status_filter)

    leaves = leaves.order_by('-expected_from')
    leave_types = LeaveType.objects.all()

    # Role-based template rendering for leave history
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        template_name = 'hr_de/leave_history_head.html'
    else:
        template_name = 'hr_de/leave_history.html'

    return render(request, template_name, {
        'leaves': leaves,
        'leave_types': leave_types,
        'query': query,
        'selected_type': leave_type_id,
    })

@login_required
def leave_edit(request, pk):
    leave = get_object_or_404(Leave, pk=pk)
    employees = Employee.objects.filter(is_active=True)

    # Restrict employees for department head
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        dept = request.user.role.department
        employees = employees.filter(department=dept) if dept else employees.none()

    leave_types = LeaveType.objects.all()

    if request.method == "POST":
        leave.employee_id = request.POST.get("employee")
        leave.leave_type_id = request.POST.get("leave_type")
        leave.expected_from = request.POST.get("expected_from")
        leave.expected_to = request.POST.get("expected_to")
        leave.reported_to = request.POST.get("reported_to")
        leave.actual_from = request.POST.get("actual_from") or None
        leave.actual_to = request.POST.get("actual_to") or None
        leave.rejoined_on = request.POST.get("rejoined_on") or None
        leave.reason = request.POST.get("reason")
  
        # Handle file upload (replace if new one uploaded)
        if request.FILES.get("leave_application"):
            leave.leave_application = request.FILES.get("leave_application")

        leave.save()
        return redirect('leave_history')

    # Role-based template rendering
    if hasattr(request.user, 'role') and request.user.role.role == 'Head':
        template_name = 'hr_de/add_leave_head.html'
    else:
        template_name = 'hr_de/add_leave.html'

    return render(request, template_name, {
        'leave': leave,
        'employees': employees,
        'leave_types': leave_types
    })


def employee_leave_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    leaves = Leave.objects.filter(employee=employee).order_by('-expected_from')
    total_leave_days = sum(
        (leave.actual_to - leave.actual_from).days + 1 for leave in leaves
        if leave.actual_from and leave.actual_to
    )
    total_leaves = Leave.objects.filter(employee=employee).count()
    sick_leaves = leaves.filter(leave_type__name='Sick').count()
    annual_leaves = leaves.filter(leave_type__name='Annual').count()
    return render(request, 'hr_de/employee_leave_detail.html', {
        'employee': employee,
        'leaves': leaves,
        'total_leave_days': total_leave_days,
        'total_leaves': total_leaves,
        'sick_leaves': sick_leaves,
        'annual_leaves': annual_leaves,
    })


@login_required
def home(request):
    from datetime import date
    today = date.today()
    total_employees = Employee.objects.count()
    pending_leaves = Leave.objects.filter(
        status__in=['Pending', 'Head_Approved', 'MD_Approved']
    ).count()
    approved_leaves = Leave.objects.filter(status='Approved').count()
    recent_leaves = Leave.objects.select_related('employee', 'leave_type').order_by('-created_at')[:5]

    # Active employees celebrating a birthday today
    birthdays_today = Employee.objects.filter(
        is_active=True, dob__month=today.month, dob__day=today.day
    ).select_related('department').order_by('emp_name')

    # Performance — this month's reviews
    reviews_this_month = MonthlyReview.objects.filter(month=today.month, year=today.year)
    reviews_count = reviews_this_month.count()
    reviews_flagged = reviews_this_month.filter(needs_attention=True).count()

    return render(request, 'hr_de/home.html', {
        'total_employees': total_employees,
        'pending_leaves': pending_leaves,
        'approved_leaves': approved_leaves,
        'recent_leaves': recent_leaves,
        'today': today,
        'birthdays_today': birthdays_today,
        'reviews_count': reviews_count,
        'reviews_flagged': reviews_flagged,
    })

@login_required
def employee_home(request):
    user_role = getattr(request.user, 'role', None)
    dept = user_role.department if user_role else None

    from django.db.models import Count, Q as DQ
    employees = Employee.objects.filter(department=dept).annotate(
        total_leaves=Count('leave'),
        pending_leaves=Count('leave', filter=DQ(leave__status='Pending')),
        approved_leaves=Count('leave', filter=DQ(leave__status='Approved')),
        rejected_leaves=Count('leave', filter=DQ(leave__status='Rejected')),
    ).order_by('emp_name') if dept else Employee.objects.none()

    pending_for_head = Leave.objects.filter(
        employee__department=dept, status='Pending'
    ).select_related('employee', 'leave_type').order_by('-created_at') if dept else Leave.objects.none()

    recent_approved = Leave.objects.filter(
        employee__department=dept, status='Approved'
    ).select_related('employee', 'leave_type').order_by('-created_at')[:10] if dept else Leave.objects.none()

    leave_types = LeaveType.objects.all()
    is_support_staff = dept.is_support_staff if dept else False

    return render(request, 'hr_de/employee_home.html', {
        'employees': employees,
        'dept': dept,
        'dept_name': dept.name if dept else '',
        'pending_for_head': pending_for_head,
        'recent_approved': recent_approved,
        'leave_types': leave_types,
        'is_support_staff': is_support_staff,
        'total_employees': employees.count(),
        'total_pending': pending_for_head.count(),
    })

@login_required
def org_chart(request):
    try:
        user_role = request.user.role.role
    except Exception:
        user_role = None
    if user_role not in ['HR', 'Manager', 'MD']:
        return render(request, 'hr_de/unauthorized.html', status=403)

    departments = Department.objects.all().order_by('name')
    chart_data = []
    for dept in departments:
        # The HR dept's head is stored as role='HR' (scoped to the dept); include it.
        head_roles = Role.objects.filter(
            department=dept, role__in=['Head', 'HR']
        ).select_related('user')
        heads = []
        for hr in head_roles:
            try:
                emp = hr.user.employee
            except Exception:
                emp = None
            heads.append({'user': hr.user, 'employee': emp})

        all_employees = Employee.objects.filter(department=dept).select_related('user')
        head_emp_pks = {h['employee'].pk for h in heads if h['employee']}
        regular = [e for e in all_employees if e.pk not in head_emp_pks]

        name = (dept.name or '').strip().upper()
        chart_data.append({
            'department': dept,
            'heads': heads,
            'employees': regular,
            'total': all_employees.count(),
            'is_shop': name.startswith('SHOP') or name.startswith('STORE'),
        })

    # Split shops/stores into their own section
    shop_data = [c for c in chart_data if c['is_shop']]
    chart_data = [c for c in chart_data if not c['is_shop']]

    # Company-level management only (dept=None) — dept-scoped HR heads show under their dept
    top_mgmt = Role.objects.filter(
        role__in=['MD', 'HR'], department__isnull=True
    ).select_related('user')
    return render(request, 'hr_de/org_chart.html', {
        'chart_data': chart_data,
        'shop_data': shop_data,
        'top_mgmt': top_mgmt,
        'can_assign_head': user_role == 'HR',
    })


@login_required
def assign_head(request):
    """AJAX: drag an employee onto a department head zone to promote them."""
    import json
    try:
        caller_role = request.user.role.role
    except Exception:
        caller_role = None
    if caller_role != 'HR':
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    try:
        data       = json.loads(request.body)
        emp_pk     = int(data['employee_pk'])
        dept_pk    = int(data['department_pk'])
    except (KeyError, ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Invalid payload.'}, status=400)

    employee   = get_object_or_404(Employee, pk=emp_pk)
    department = get_object_or_404(Department, pk=dept_pk)

    if not employee.user:
        return JsonResponse({
            'ok': False,
            'error': f'"{employee.emp_name}" has no login account. '
                     'Assign a username via Edit Employee first.'
        })

    # Assign / update the head role (head of the HR dept becomes full HR)
    role_value = _head_role_for(department)
    Role.objects.update_or_create(
        user=employee.user,
        defaults={'role': role_value, 'department': department}
    )

    # Move employee to the target department if they belong to a different one
    if employee.department_id != dept_pk:
        employee.department = department
        employee.save(update_fields=['department'])

    label = 'HR' if role_value == 'HR' else 'Head'
    return JsonResponse({
        'ok': True,
        'message': f'{employee.emp_name} is now {label} of {department.name}.'
    })


def settings(request):
    return render(request, 'hr_de/settings.html', {})


# ── Employee self-service ────────────────────────────────────────────────────

@login_required
def employee_dashboard(request):
    """Landing dashboard for a regular (self-service) employee — greeting + feature tiles."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    leaves        = Leave.objects.filter(employee=employee)
    advances      = AdvanceSalary.objects.filter(employee=employee)
    passport_reqs = PassportRequest.objects.filter(employee=employee)

    stats = {
        'pending_leaves':   leaves.filter(status__in=['Pending', 'Head_Approved', 'MD_Approved']).count(),
        'approved_leaves':  leaves.filter(status='Approved').count(),
        'pending_advances': advances.filter(status__in=['Pending', 'Head_Approved']).count(),
        'active_advances':  advances.filter(status='Approved').count(),
        'pending_passport': passport_reqs.filter(status='Pending').count(),
        'outstanding_passport': passport_reqs.filter(status='Approved').count(),
    }

    from datetime import date
    _today = date.today()
    is_birthday = bool(employee.dob) and (employee.dob.month, employee.dob.day) == (_today.month, _today.day)

    return render(request, 'hr_de/employee_dashboard.html', {
        'employee': employee,
        'stats': stats,
        'is_birthday': is_birthday,
    })


@login_required
def my_leaves(request):
    """Personal leave dashboard for regular employees."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    leaves = Leave.objects.filter(employee=employee).select_related('leave_type').order_by('-created_at')
    leave_types = LeaveType.objects.all()

    total  = leaves.count()
    pending  = leaves.filter(status='Pending').count()
    approved = leaves.filter(status='Approved').count()
    rejected = leaves.filter(status='Rejected').count()

    return render(request, 'hr_de/my_leaves.html', {
        'employee': employee,
        'leaves': leaves,
        'leave_types': leave_types,
        'total': total,
        'pending': pending,
        'approved': approved,
        'rejected': rejected,
    })


@login_required
def submit_my_leave(request):
    """Employee submits their own leave request."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    leave_types = LeaveType.objects.all()

    if request.method == 'POST':
        expected_from = request.POST.get('expected_from')
        expected_to   = request.POST.get('expected_to')
        if expected_from and expected_to:
            from datetime import date
            days = (
                date.fromisoformat(expected_to) - date.fromisoformat(expected_from)
            ).days + 1
        else:
            days = 0

        Leave.objects.create(
            employee=employee,
            leave_type_id=request.POST.get('leave_type'),
            expected_from=expected_from,
            expected_to=expected_to,
            reported_to=request.POST.get('reported_to', ''),
            reason=request.POST.get('reason', ''),
            days=days,
            leave_application=request.FILES.get('leave_application'),
            status='Pending',
        )
        messages.success(request, "Leave request submitted successfully.")
        return redirect('my_leaves')

    return render(request, 'hr_de/submit_leave.html', {
        'employee': employee,
        'leave_types': leave_types,
    })


# ── Head: apply leave on behalf of any employee in their department ──────────

@login_required
def apply_leave_behalf(request, emp_pk):
    """Head applies leave for any employee under their department.
    Auto-approves the Head stage; support-staff depts also skip MD."""
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if user_role.role != 'Head':
        return render(request, 'hr_de/unauthorized.html', status=403)

    dept = user_role.department
    employee = get_object_or_404(Employee, pk=emp_pk, department=dept)
    leave_types = LeaveType.objects.all()

    if request.method == 'POST':
        from datetime import date as _date
        expected_from_str = request.POST.get('expected_from')
        expected_to_str   = request.POST.get('expected_to')
        rejoined_on_str   = request.POST.get('rejoined_on', '').strip()

        expected_from = _date.fromisoformat(expected_from_str) if expected_from_str else None
        expected_to   = _date.fromisoformat(expected_to_str)   if expected_to_str   else None

        if expected_from and expected_to:
            days = (expected_to - expected_from).days + 1
        else:
            days = 0

        # Rejoining date defaults to expected_to; actual_days recalculated from it
        if rejoined_on_str:
            rejoined_on  = _date.fromisoformat(rejoined_on_str)
            actual_days  = (rejoined_on - expected_from).days + 1 if expected_from else days
        else:
            rejoined_on  = expected_to
            actual_days  = days

        # Head already approved — determine next stage based on dept type
        is_support = dept.is_support_staff if dept else False
        if is_support:
            # Support staff: skip Head + MD → goes straight to HR
            status = 'MD_Approved'
        else:
            # Regular dept: skip Head → goes to MD next
            status = 'Head_Approved'

        Leave.objects.create(
            employee=employee,
            leave_type_id=request.POST.get('leave_type'),
            expected_from=expected_from,
            expected_to=expected_to,
            reported_to=request.POST.get('reported_to', f"{request.user.get_full_name() or request.user.username} (Head)"),
            reason=request.POST.get('reason', ''),
            days=days,
            actual_days=actual_days,
            rejoined_on=rejoined_on,
            leave_application=request.FILES.get('leave_application'),
            status=status,
            head_approved_by=request.user,
            head_approved_at=timezone.now(),
        )
        messages.success(request, f"Leave applied for {employee.emp_name} and auto-approved at Head stage.")
        return redirect('employee_home')

    return render(request, 'hr_de/apply_leave_behalf.html', {
        'employee': employee,
        'leave_types': leave_types,
        'dept': dept,
    })


# ── Head: update rejoining date on a dept employee's leave ───────────────────

@login_required
def set_rejoining_date(request, leave_pk):
    """Head sets/updates the actual rejoining date for a leave in their dept.
    Recalculates actual_days = (rejoined_on - expected_from).days + 1."""
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if user_role.role != 'Head':
        return render(request, 'hr_de/unauthorized.html', status=403)

    leave = get_object_or_404(Leave, pk=leave_pk, employee__department=user_role.department)

    if request.method == 'POST':
        from datetime import date as _date
        rejoined_str = request.POST.get('rejoined_on', '').strip()
        if rejoined_str and leave.expected_from:
            rejoined_on = _date.fromisoformat(rejoined_str)
            actual_days = (rejoined_on - leave.expected_from).days + 1
            leave.rejoined_on = rejoined_on
            leave.actual_days = actual_days
            leave.save()
            messages.success(request, f"Rejoining date updated for {leave.employee.emp_name}. Actual days: {actual_days}.")
        else:
            messages.error(request, "Invalid rejoining date.")
        return redirect('leave_history')

    return render(request, 'hr_de/set_rejoining_date.html', {'leave': leave})


# ── HR: set / edit salary structure for an employee ─────────────────────────

@login_required
def salary_structure(request, emp_pk):
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)
    if user_role.role != 'HR':
        return render(request, 'hr_de/unauthorized.html', status=403)

    employee = get_object_or_404(Employee, pk=emp_pk)

    # Get or create the salary structure record
    structure, _ = SalaryStructure.objects.get_or_create(employee=employee)

    if request.method == 'POST':
        def _float(key):
            try:
                return float(request.POST.get(key, 0) or 0)
            except ValueError:
                return 0.0

        structure.basic     = _float('basic')
        structure.hra       = _float('hra')
        structure.transport = _float('transport')
        structure.fuel      = _float('fuel')
        structure.others    = _float('others')
        structure.updated_by = request.user
        structure.save()
        messages.success(request, f"Salary structure updated for {employee.emp_name}.")
        return redirect('employee_detail', pk=emp_pk)

    return render(request, 'hr_de/salary_structure.html', {
        'employee': employee,
        'structure': structure,
    })


# ── PAYROLL VIEWS ────────────────────────────────────────────────────────────

def _hr_only(request):
    try:
        return request.user.role.role == 'HR'
    except Exception:
        return False


def _hr_or_md(request):
    try:
        return request.user.role.role in ('HR', 'MD')
    except Exception:
        return False


@login_required
def payroll_list(request):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    runs = PayrollRun.objects.prefetch_related('entries').all()
    return render(request, 'hr_de/payroll/payroll_list.html', {
        'runs': runs,
        'draft_count': runs.filter(status='Draft').count(),
        'confirmed_count': runs.filter(status='Confirmed').count(),
    })


@login_required
def payroll_create(request):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        month = int(request.POST.get('month', 0))
        year  = int(request.POST.get('year',  0))
        notes = request.POST.get('notes', '')

        if not (1 <= month <= 12) or year < 2000:
            messages.error(request, "Please choose a valid month and year.")
            return redirect('payroll_create')

        if PayrollRun.objects.filter(month=month, year=year).exists():
            messages.error(request, f"Payroll for {month}/{year} already exists.")
            return redirect('payroll_list')

        # Block future months — attendance for a month that hasn't started yet
        # can't be complete, so payroll must not run for it.
        import datetime as _dt
        _today = timezone.localdate()
        if _dt.date(year, month, 1) > _dt.date(_today.year, _today.month, 1):
            month_name = dict(PayrollRun.MONTH_CHOICES).get(month, month)
            messages.error(
                request,
                f"Cannot run payroll for {month_name} {year} — that month hasn't "
                f"started yet. Attendance can only be completed once the month begins."
            )
            return redirect('payroll_create')

        # Attendance must be complete for the month before payroll can run
        gaps = _attendance_incomplete(year, month)
        if gaps:
            total_missing = sum(m for _, m in gaps)
            preview = ', '.join(
                f"{e.emp_name} ({m} day{'s' if m != 1 else ''})" for e, m in gaps[:6]
            )
            more = '' if len(gaps) <= 6 else f", +{len(gaps) - 6} more"
            month_name = dict(PayrollRun.MONTH_CHOICES).get(month, month)
            messages.error(
                request,
                f"Cannot run payroll — attendance for {month_name} {year} is incomplete. "
                f"{len(gaps)} employee(s) have {total_missing} unfilled day(s): {preview}{more}. "
                f"Please complete the Attendance Sheet for every employee first."
            )
            return redirect('payroll_create')

        run = PayrollRun.objects.create(
            month=month, year=year, notes=notes, created_by=request.user
        )

        # Auto-populate one entry per active employee that has a salary structure
        employees = Employee.objects.filter(is_active=True).select_related('salary_structure')
        created = 0
        for emp in employees:
            try:
                s = emp.salary_structure
            except Exception:
                s = None

            # Check for an active approved advance salary with remaining balance
            advance_ded, advance_obj = _pick_active_advance(emp)

            gross = s.total if s else 0
            absent, half, att_ded = _attendance_deduction_for(emp, year, month, gross)
            entry = PayrollEntry(
                payroll_run=run,
                employee=emp,
                basic=s.basic if s else 0,
                hra=s.hra if s else 0,
                transport=s.transport if s else 0,
                fuel=s.fuel if s else 0,
                others=s.others if s else 0,
                advance_deduction=advance_ded,
                advance_salary=advance_obj,
                absent_days=absent,
                half_days=half,
                attendance_deduction=att_ded,
            )
            entry.compute_and_save()
            created += 1

        messages.success(request, f"Payroll run created with {created} employee entries.")
        return redirect('payroll_detail', pk=run.pk)

    from datetime import date
    today = date.today()
    return render(request, 'hr_de/payroll/payroll_create.html', {
        'current_month': today.month,
        'current_year':  today.year,
        'month_choices': PayrollRun.MONTH_CHOICES,
    })


def _attendance_deduction_for(employee, year, month, gross):
    """Return (absent_days, half_days, deduction) for an employee in a month.

    Per-day rate = gross salary / total days in the month. Each absent day deducts
    one full day; each half day deducts half a day.
    """
    import calendar
    import datetime as dt
    num_days = calendar.monthrange(year, month)[1]
    first = dt.date(year, month, 1)
    last  = dt.date(year, month, num_days)
    qs = Attendance.objects.filter(employee=employee, date__range=(first, last))
    absent = qs.filter(status='absent').count()
    half   = qs.filter(status='half_day').count()
    per_day = (gross / num_days) if (num_days and gross) else 0
    deduction = round(per_day * absent + (per_day / 2.0) * half, 2)
    return absent, half, deduction


def _attendance_incomplete(year, month):
    """Return [(employee, missing_days), ...] for active employees with blank
    attendance days that must be filled before payroll can run.

    A day counts as filled if it has an Attendance record, is covered by an
    approved leave, or is a Sunday (auto week-off). Days before an employee's
    joining date and days still in the future are not required.
    """
    import calendar
    import datetime as dt
    today    = timezone.localdate()
    num_days = calendar.monthrange(year, month)[1]
    first    = dt.date(year, month, 1)
    last     = dt.date(year, month, num_days)
    cutoff   = min(last, today)

    employees = list(Employee.objects.filter(is_active=True))
    emp_ids   = [e.id for e in employees]

    have = set()
    for eid, d in Attendance.objects.filter(
        employee_id__in=emp_ids, date__range=(first, last)
    ).values_list('employee_id', 'date'):
        have.add((eid, d.day))

    leave_days = set()
    for lv in Leave.objects.filter(employee_id__in=emp_ids, status='Approved'):
        f = lv.actual_from or lv.expected_from
        t = lv.actual_to or lv.expected_to
        if not f or not t:
            continue
        cur = max(f, first)
        end = min(t, last)
        while cur <= end:
            leave_days.add((lv.employee_id, cur.day))
            cur += dt.timedelta(days=1)

    gaps = []
    for emp in employees:
        missing = 0
        for d in range(1, num_days + 1):
            dobj = dt.date(year, month, d)
            if dobj > cutoff:
                continue                              # future day — not yet required
            if emp.joining_date and dobj < emp.joining_date:
                continue                              # before joining — excused
            if dobj.weekday() == 6:
                continue                              # Sunday auto week-off
            if (emp.id, d) in have or (emp.id, d) in leave_days:
                continue
            missing += 1
        if missing:
            gaps.append((emp, missing))
    return gaps


def _sync_attendance_deductions(run):
    """Refresh attendance deductions on a DRAFT run from current attendance records."""
    if run.status != 'Draft':
        return
    for entry in run.entries.select_related('employee'):
        absent, half, ded = _attendance_deduction_for(
            entry.employee, run.year, run.month, entry.gross_salary or 0
        )
        if (entry.absent_days != absent or entry.half_days != half
                or round(entry.attendance_deduction, 2) != round(ded, 2)):
            entry.absent_days          = absent
            entry.half_days            = half
            entry.attendance_deduction = ded
            entry.compute_and_save()


def _pick_active_advance(employee):
    """Return (deduction_amount, advance_obj) for the employee's oldest active advance."""
    for adv in employee.advance_salaries.filter(status='Approved').order_by('created_at'):
        remaining = adv.remaining_amount
        monthly   = adv.effective_monthly_deduction
        if remaining > 0 and monthly and monthly > 0:
            return round(min(monthly, remaining), 2), adv
    return 0.0, None


def _sync_advance_deductions(run):
    """Refresh advance deductions on a DRAFT run from current approved advances.

    Confirmed runs are locked snapshots and are never modified. Because
    remaining_amount only counts *confirmed* deductions, re-syncing a draft is
    idempotent and never double-counts.
    """
    if run.status != 'Draft':
        return
    for entry in run.entries.select_related('employee'):
        advance_ded, advance_obj = _pick_active_advance(entry.employee)
        advance_obj_pk = advance_obj.pk if advance_obj else None
        if round(entry.advance_deduction, 2) != advance_ded or entry.advance_salary_id != advance_obj_pk:
            entry.advance_deduction = advance_ded
            entry.advance_salary    = advance_obj
            entry.compute_and_save()


@login_required
def payroll_detail(request, pk):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    run = get_object_or_404(PayrollRun, pk=pk)

    # Draft runs auto-refresh advance + attendance deductions so later edits appear
    _sync_advance_deductions(run)
    _sync_attendance_deductions(run)

    entries = run.entries.select_related(
        'employee', 'employee__department', 'employee__mol', 'advance_salary'
    ).order_by('employee__mol__mol', 'employee__emp_name')

    # Pre-group by MOL with subtotals so the template can render them
    mol_groups = []
    for entry in entries:
        mol = entry.employee.mol
        mol_key = mol.pk if mol else None
        mol_label = mol.mol if mol else 'No MOL Assigned'
        if mol_groups and mol_groups[-1]['mol_key'] == mol_key:
            mol_groups[-1]['entries'].append(entry)
            mol_groups[-1]['total_gross'] = round(mol_groups[-1]['total_gross'] + entry.gross_salary, 2)
            mol_groups[-1]['total_net']   = round(mol_groups[-1]['total_net']   + entry.net_salary, 2)
        else:
            mol_groups.append({
                'mol_key':    mol_key,
                'mol_label':  mol_label,
                'entries':    [entry],
                'total_gross': round(entry.gross_salary, 2),
                'total_net':   round(entry.net_salary, 2),
            })

    return render(request, 'hr_de/payroll/payroll_detail.html', {
        'run':        run,
        'entries':    entries,
        'mol_groups': mol_groups,
        'today':      timezone.localdate(),
    })


@login_required
def payroll_wps_export(request, pk):
    """Export selected payroll entries as a WPS SIF (.xlsx): EDR rows per employee
    followed by a single SCR footer row. No headers — the bank's exact layout."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    run = get_object_or_404(PayrollRun, pk=pk)
    if request.method != 'POST':
        return redirect('payroll_detail', pk=pk)

    import calendar
    import datetime as dt
    from openpyxl import Workbook
    from django.http import HttpResponse

    entry_ids = request.POST.getlist('entries')
    entries = list(
        run.entries.filter(pk__in=entry_ids)
        .select_related('employee', 'employee__mol')
        .order_by('employee__emp_name')
    )
    if not entries:
        messages.error(request, 'Select at least one employee to export.')
        return redirect('payroll_detail', pk=pk)

    # WPS files are per establishment (MOL) — all selected must share one MOL
    mol_ids = {e.employee.mol_id for e in entries}
    if len(mol_ids) > 1:
        messages.error(request, 'WPS export must be for a single MOL — select employees from one company only.')
        return redirect('payroll_detail', pk=pk)
    mol = entries[0].employee.mol
    if mol is None:
        messages.error(request, 'The selected employees have no MOL assigned.')
        return redirect('payroll_detail', pk=pk)

    # Payroll period
    num_days  = calendar.monthrange(run.year, run.month)[1]
    date_from = dt.date(run.year, run.month, 1).strftime('%Y-%m-%d')
    date_to   = dt.date(run.year, run.month, num_days).strftime('%Y-%m-%d')
    salary_month = f"{run.month:02d}{run.year}"          # e.g. 062026

    proc_date = (request.POST.get('processing_date') or '').strip() or dt.date.today().strftime('%Y-%m-%d')
    proc_time = (request.POST.get('processing_time') or '').strip()

    def _t(v):
        return '' if v is None else str(v)

    wb = Workbook()
    ws = wb.active
    ws.title = 'SIF'
    TEXT = '@'   # keep leading zeros on IBAN / labour card / routing / codes

    r = 0
    total_net = 0.0
    for e in entries:
        emp = e.employee
        net = round(e.net_salary or 0, 2)
        total_net += net
        r += 1
        ws.append([
            'EDR',
            _t(emp.labour_card_number),
            emp.emp_name or '',
            _t(emp.routing_code),
            _t(emp.iban),
            date_from,
            date_to,
            num_days,
            net,
            0.0,
            0,
        ])
        for col in (2, 4, 5, 6, 7):          # text-format the id/date columns
            ws.cell(row=r, column=col).number_format = TEXT
        ws.cell(row=r, column=10).number_format = '0.00'   # variable → 0.00

    # SCR footer
    r += 1
    ws.append([
        'SCR',
        _t(mol.company_code),
        '',
        _t(mol.wps_number),
        proc_date,
        _t(proc_time),
        salary_month,
        len(entries),
        round(total_net, 2),
        'AED',
        _t(mol.iban),
    ])
    for col in (2, 4, 5, 6, 7, 11):
        ws.cell(row=r, column=col).number_format = TEXT

    # Sensible column widths
    for i, w in enumerate([8, 18, 34, 14, 26, 12, 12, 8, 12, 10, 26], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    safe_mol = ''.join(c for c in (mol.mol or 'MOL') if c.isalnum() or c in ' _-').strip() or 'MOL'
    filename = f"WPS_SIF_{safe_mol}_{salary_month}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def payroll_entry_update(request, entry_pk):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    entry = get_object_or_404(PayrollEntry, pk=entry_pk)

    if entry.payroll_run.status == 'Confirmed':
        messages.error(request, "Cannot edit a confirmed payroll run.")
        return redirect('payroll_detail', pk=entry.payroll_run.pk)

    if request.method == 'POST':
        def _f(k):
            try: return float(request.POST.get(k, 0) or 0)
            except ValueError: return 0.0

        entry.bonus           = _f('bonus')
        entry.overtime_pay    = _f('overtime_pay')
        entry.other_additions = _f('other_additions')
        entry.loan_deduction  = _f('loan_deduction')
        entry.other_deductions= _f('other_deductions')
        entry.notes           = request.POST.get('notes', '')
        entry.compute_and_save()
        messages.success(request, f"Updated payroll entry for {entry.employee.emp_name}.")
        return redirect('payroll_detail', pk=entry.payroll_run.pk)

    # GET: refresh the attendance deduction from the current sheet so it displays live
    if entry.payroll_run.status == 'Draft':
        absent, half, ded = _attendance_deduction_for(
            entry.employee, entry.payroll_run.year, entry.payroll_run.month, entry.gross_salary or 0
        )
        if (entry.absent_days != absent or entry.half_days != half
                or round(entry.attendance_deduction, 2) != round(ded, 2)):
            entry.absent_days          = absent
            entry.half_days            = half
            entry.attendance_deduction = ded
            entry.compute_and_save()

    return render(request, 'hr_de/payroll/payroll_entry_edit.html', {'entry': entry})


@login_required
def payroll_delete(request, pk):
    """Delete an entire Draft payroll run."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    run = get_object_or_404(PayrollRun, pk=pk)
    if run.status == 'Confirmed':
        messages.error(request, 'Cannot delete a confirmed payroll run.')
        return redirect('payroll_detail', pk=pk)
    if request.method == 'POST':
        label = str(run)
        run.delete()
        messages.success(request, f'{label} has been deleted.')
        return redirect('payroll_list')
    return render(request, 'hr_de/payroll/payroll_delete_confirm.html', {'run': run})


@login_required
def payroll_entry_remove(request, entry_pk):
    """Remove a single employee entry from a Draft payroll run."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    entry = get_object_or_404(PayrollEntry, pk=entry_pk)
    if entry.payroll_run.status == 'Confirmed':
        messages.error(request, 'Cannot remove entries from a confirmed payroll run.')
        return redirect('payroll_detail', pk=entry.payroll_run.pk)
    if request.method == 'POST':
        run_pk = entry.payroll_run.pk
        emp_name = entry.employee.emp_name
        entry.delete()
        messages.success(request, f'Removed {emp_name} from the payroll run.')
        return redirect('payroll_detail', pk=run_pk)
    return redirect('payroll_detail', pk=entry.payroll_run.pk)


@login_required
def payroll_confirm(request, pk):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    run = get_object_or_404(PayrollRun, pk=pk)
    if request.method == 'POST':
        if run.status == 'Confirmed':
            messages.info(request, "Already confirmed.")
        else:
            run.status = 'Confirmed'
            run.confirmed_at = timezone.now()
            run.save()
            messages.success(request, f"{run} has been confirmed and locked.")
    return redirect('payroll_detail', pk=pk)


@login_required
def payslip(request, entry_pk):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    entry = get_object_or_404(
        PayrollEntry.objects.select_related('employee', 'employee__department', 'payroll_run'),
        pk=entry_pk
    )
    return render(request, 'hr_de/payroll/payslip.html', {'entry': entry})


############ ADVANCE SALARY ############

@login_required
def advance_apply(request):
    """Submit a new advance salary application."""
    try:
        user_role = request.user.role.role
    except Exception:
        user_role = None

    # Determine which employees are selectable
    if user_role == 'HR':
        employees = Employee.objects.filter(is_active=True).order_by('emp_name')
    elif user_role == 'Head':
        dept = request.user.role.department
        employees = Employee.objects.filter(is_active=True, department=dept).order_by('emp_name') if dept else Employee.objects.none()
    else:
        # Regular employee — can only apply for themselves
        try:
            employees = [request.user.employee]
        except Exception:
            return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        emp_pk      = request.POST.get('employee')
        amount_raw  = request.POST.get('amount', '').strip()
        reason      = request.POST.get('reason', '').strip()
        monthly_raw = request.POST.get('requested_monthly_deduction', '').strip()

        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'Please enter a valid positive amount.')
            return redirect('advance_apply')

        try:
            req_monthly = float(monthly_raw)
            if req_monthly <= 0 or req_monthly > amount:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'Monthly deduction must be a positive value and not exceed the advance amount.')
            return redirect('advance_apply')

        if not reason:
            messages.error(request, 'Reason is required.')
            return redirect('advance_apply')

        if user_role in ('HR', 'Head'):
            employee = get_object_or_404(Employee, pk=emp_pk)
        else:
            employee = request.user.employee

        # Head applying on behalf → auto-approve Head stage (keeps requested amount)
        if user_role == 'Head':
            initial_status  = 'Head_Approved'
            head_approved_by = request.user
            head_approved_at = timezone.now()
            approved_monthly = req_monthly
        else:
            initial_status   = 'Pending'
            head_approved_by = None
            head_approved_at = None
            approved_monthly = None

        AdvanceSalary.objects.create(
            employee                    = employee,
            amount                      = amount,
            reason                      = reason,
            requested_monthly_deduction = req_monthly,
            approved_monthly_deduction  = approved_monthly,
            status                      = initial_status,
            head_approved_by            = head_approved_by,
            head_approved_at            = head_approved_at,
            applied_by                  = request.user,
        )
        messages.success(request, f'Advance salary application submitted for {employee.emp_name}.')
        if user_role == 'HR':
            return redirect('advance_list')
        elif user_role == 'Head':
            return redirect('advance_pending')
        else:
            return redirect('my_advances')

    return render(request, 'hr_de/advance_salary/apply.html', {'employees': employees, 'user_role': user_role})


@login_required
def advance_pending(request):
    """Pending queue — Head sees Pending for their dept; HR sees Head_Approved."""
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if user_role.role == 'Head':
        dept    = user_role.department
        advances = AdvanceSalary.objects.filter(
            status='Pending', employee__department=dept
        ).select_related('employee', 'employee__department') if dept else AdvanceSalary.objects.none()
    elif user_role.role == 'HR':
        advances = AdvanceSalary.objects.filter(
            status='Head_Approved'
        ).select_related('employee', 'employee__department', 'head_approved_by')
    else:
        return render(request, 'hr_de/unauthorized.html', status=403)

    return render(request, 'hr_de/advance_salary/pending.html', {
        'advances':  advances,
        'user_role': user_role,
    })


@login_required
def advance_approve(request, pk):
    """Approve or reject a single advance salary application."""
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    advance = get_object_or_404(AdvanceSalary, pk=pk)

    if request.method == 'POST':
        action           = request.POST.get('action')
        rejection_reason = request.POST.get('rejection_reason', '').strip()

        def _approved_amount(default):
            try:
                val = float(request.POST.get('approved_monthly_deduction', '') or default)
                return round(val, 2) if val > 0 else round(float(default), 2)
            except (ValueError, TypeError):
                return round(float(default), 2)

        if user_role.role == 'Head':
            if advance.status != 'Pending':
                messages.error(request, 'This application is no longer pending for Head approval.')
                return redirect('advance_pending')
            if advance.employee.department != user_role.department:
                return render(request, 'hr_de/unauthorized.html', status=403)
            if action == 'approve':
                approved_amt = _approved_amount(advance.requested_monthly_deduction)
                advance.status           = 'Head_Approved'
                advance.head_approved_by = request.user
                advance.head_approved_at = timezone.now()
                advance.save()
                # Set + log the monthly deduction (logs only if it differs)
                advance.set_monthly_deduction(approved_amt, request.user, 'Head', note='Set at Head approval')
            elif action == 'reject':
                advance.status           = 'Rejected'
                advance.head_approved_by = request.user
                advance.rejection_reason = rejection_reason
                advance.save()

        elif user_role.role == 'HR':
            if advance.status != 'Head_Approved':
                messages.error(request, 'This application must be approved by the Head first.')
                return redirect('advance_pending')
            if action == 'approve':
                approved_amt = _approved_amount(advance.effective_monthly_deduction)
                advance.status        = 'Approved'
                advance.hr_approved_by = request.user
                advance.hr_approved_at = timezone.now()
                advance.save()
                advance.set_monthly_deduction(approved_amt, request.user, 'HR', note='Set at HR final approval')
            elif action == 'reject':
                advance.status        = 'Rejected'
                advance.hr_approved_by = request.user
                advance.rejection_reason = rejection_reason
                advance.save()
        else:
            return render(request, 'hr_de/unauthorized.html', status=403)

        messages.success(request, f'Application for {advance.employee.emp_name} has been {advance.status.lower().replace("_", " ")}.')
        return redirect('advance_pending')

    return render(request, 'hr_de/advance_salary/approve.html', {
        'advance':   advance,
        'user_role': user_role,
    })


@login_required
def advance_edit_deduction(request, pk):
    """Ongoing edit of the monthly deduction — HR (any) or Head (their dept)."""
    try:
        user_role = request.user.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    advance = get_object_or_404(
        AdvanceSalary.objects.select_related('employee', 'employee__department'), pk=pk
    )

    # Permission: HR any; Head only their own department
    if user_role.role == 'HR':
        pass
    elif user_role.role == 'Head':
        if advance.employee.department != user_role.department:
            return render(request, 'hr_de/unauthorized.html', status=403)
    else:
        return render(request, 'hr_de/unauthorized.html', status=403)

    # Cannot edit a rejected application
    if advance.status == 'Rejected':
        messages.error(request, 'Cannot edit the deduction of a rejected application.')
        return redirect('advance_list')

    if request.method == 'POST':
        note = request.POST.get('note', '').strip()
        try:
            new_amt = float(request.POST.get('approved_monthly_deduction', ''))
            if new_amt <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'Enter a valid positive monthly deduction amount.')
            return redirect('advance_edit_deduction', pk=pk)

        if new_amt > advance.remaining_amount and advance.status == 'Approved':
            messages.error(request, f'Monthly deduction cannot exceed the remaining balance (AED {advance.remaining_amount:.2f}).')
            return redirect('advance_edit_deduction', pk=pk)

        advance.set_monthly_deduction(new_amt, request.user, user_role.role, note=note)
        messages.success(request, f'Monthly deduction for {advance.employee.emp_name} updated to AED {new_amt:.2f}.')
        return redirect('advance_edit_deduction', pk=pk)

    changes = advance.deduction_changes.select_related('changed_by').all()
    return render(request, 'hr_de/advance_salary/edit_deduction.html', {
        'advance':   advance,
        'user_role': user_role,
        'changes':   changes,
    })


@login_required
def advance_list(request):
    """Full history — HR sees all; Head sees their own department."""
    try:
        role_obj  = request.user.role
        user_role = role_obj.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if user_role not in ('HR', 'Head'):
        return render(request, 'hr_de/unauthorized.html', status=403)

    status_filter = request.GET.get('status', '')
    query         = request.GET.get('q', '')

    advances = AdvanceSalary.objects.select_related(
        'employee', 'employee__department', 'head_approved_by', 'hr_approved_by'
    )

    # Head is scoped to their own department
    if user_role == 'Head':
        dept = role_obj.department
        advances = advances.filter(employee__department=dept) if dept else advances.none()

    if status_filter:
        advances = advances.filter(status=status_filter)
    if query:
        advances = advances.filter(
            Q(employee__emp_name__icontains=query) |
            Q(employee__emp_id__icontains=query)
        )

    return render(request, 'hr_de/advance_salary/list.html', {
        'advances':        advances,
        'status_choices':  AdvanceSalary.STATUS_CHOICES,
        'selected_status': status_filter,
        'query':           query,
        'user_role':       user_role,
    })


@login_required
def my_advances(request):
    """Employee's own advance history."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    advances = AdvanceSalary.objects.filter(employee=employee)
    return render(request, 'hr_de/advance_salary/my_advances.html', {
        'employee': employee,
        'advances': advances,
    })


############ PASSPORT REQUESTS ############

@login_required
def passport_apply(request):
    """Submit a passport request. HR can pick any employee; others apply for self."""
    try:
        user_role = request.user.role.role
    except Exception:
        user_role = None

    if user_role == 'HR':
        employees = Employee.objects.filter(is_active=True).order_by('emp_name')
    elif user_role == 'Head':
        dept = request.user.role.department
        employees = Employee.objects.filter(is_active=True, department=dept).order_by('emp_name') if dept else Employee.objects.none()
    else:
        try:
            employees = [request.user.employee]
        except Exception:
            return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        emp_pk          = request.POST.get('employee')
        reason          = request.POST.get('reason', '').strip()
        needed_from     = _parse_date(request.POST.get('needed_from') or None)
        expected_return = _parse_date(request.POST.get('expected_return') or None)

        if not reason:
            messages.error(request, 'Reason is required.')
            return redirect('passport_apply')

        if user_role == 'HR':
            employee = get_object_or_404(Employee, pk=emp_pk)
        elif user_role == 'Head':
            # Head can only request on behalf of employees in their own department
            dept = request.user.role.department
            employee = get_object_or_404(Employee, pk=emp_pk, department=dept)
        else:
            employee = request.user.employee

        PassportRequest.objects.create(
            employee        = employee,
            reason          = reason,
            needed_from     = needed_from,
            expected_return = expected_return,
            applied_by      = request.user,
        )
        messages.success(request, f'Passport request submitted for {employee.emp_name}.')
        return redirect('passport_list' if user_role in ('HR', 'Head') else 'my_passport_requests')

    return render(request, 'hr_de/passport/apply.html', {
        'employees': employees,
        'user_role': user_role,
    })


@login_required
def passport_pending(request):
    """HR queue of pending passport requests."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    requests_qs = PassportRequest.objects.filter(status='Pending').select_related(
        'employee', 'employee__department'
    )
    return render(request, 'hr_de/passport/pending.html', {'requests': requests_qs})


@login_required
def passport_approve(request, pk):
    """HR approves or rejects a passport request."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    req = get_object_or_404(PassportRequest.objects.select_related('employee'), pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if req.status != 'Pending':
            messages.error(request, 'This request has already been processed.')
            return redirect('passport_pending')

        if action == 'approve':
            req.status      = 'Approved'
            req.approved_by = request.user
            req.approved_at = timezone.now()
            req.save()
            # Passport is now physically with the employee
            req.employee.passport_status = 'With employee'
            req.employee.save(update_fields=['passport_status'])
            messages.success(request, f'Passport request for {req.employee.emp_name} approved.')
        elif action == 'reject':
            req.status           = 'Rejected'
            req.approved_by      = request.user
            req.approved_at      = timezone.now()
            req.rejection_reason = request.POST.get('rejection_reason', '').strip()
            req.save()
            messages.success(request, f'Passport request for {req.employee.emp_name} rejected.')
        return redirect('passport_pending')

    return render(request, 'hr_de/passport/approve.html', {'req': req})


@login_required
def passport_mark_returned(request, pk):
    """HR marks an approved passport as returned to the company."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    req = get_object_or_404(PassportRequest.objects.select_related('employee'), pk=pk)
    if request.method == 'POST':
        if req.status != 'Approved':
            messages.error(request, 'Only an approved (outstanding) request can be marked returned.')
            return redirect('passport_list')
        req.status             = 'Returned'
        req.returned_at        = timezone.now()
        req.returned_marked_by = request.user
        req.save()
        req.employee.passport_status = 'With company'
        req.employee.save(update_fields=['passport_status'])
        messages.success(request, f"{req.employee.emp_name}'s passport has been returned and is back with the company.")
    # Return to wherever the action was triggered (outstanding page or full list)
    next_url = request.POST.get('next')
    if next_url == 'outstanding':
        return redirect('passport_outstanding')
    return redirect('passport_list')


@login_required
def passport_outstanding(request):
    """Passports currently OUT — any employee whose custody is 'With employee'.

    Includes employees set manually via the employee edit form (no request)
    as well as those approved through a passport request.
    """
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    from datetime import date

    employees_out = Employee.objects.filter(
        passport_status='With employee'
    ).select_related('department').order_by('emp_name')

    items = []
    for emp in employees_out:
        req = emp.passport_requests.filter(status='Approved').order_by('-approved_at').first()
        items.append({'employee': emp, 'request': req})

    return render(request, 'hr_de/passport/outstanding.html', {
        'items': items,
        'today': date.today(),
    })


@login_required
def passport_employee_return(request, emp_pk):
    """Mark an employee's passport as returned to the company.

    Flips custody back to 'With company' and closes any open approved request.
    Works whether or not a passport request exists.
    """
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    emp = get_object_or_404(Employee, pk=emp_pk)
    if request.method == 'POST':
        for req in emp.passport_requests.filter(status='Approved'):
            req.status             = 'Returned'
            req.returned_at        = timezone.now()
            req.returned_marked_by = request.user
            req.save()
        emp.passport_status = 'With company'
        emp.save(update_fields=['passport_status'])
        messages.success(request, f"{emp.emp_name}'s passport has been returned and is back with the company.")
    return redirect('passport_outstanding')


@login_required
def passport_list(request):
    """Passport request history. HR sees all; Head sees their own department."""
    try:
        role_obj  = request.user.role
        user_role = role_obj.role
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)

    if user_role not in ('HR', 'Head'):
        return render(request, 'hr_de/unauthorized.html', status=403)

    status_filter = request.GET.get('status', '')
    query         = request.GET.get('q', '')

    requests_qs = PassportRequest.objects.select_related(
        'employee', 'employee__department', 'approved_by'
    )

    # Head is scoped to their own department
    if user_role == 'Head':
        dept = role_obj.department
        requests_qs = requests_qs.filter(employee__department=dept) if dept else requests_qs.none()

    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)
    if query:
        requests_qs = requests_qs.filter(
            Q(employee__emp_name__icontains=query) |
            Q(employee__emp_id__icontains=query)
        )

    return render(request, 'hr_de/passport/list.html', {
        'requests':        requests_qs,
        'status_choices':  PassportRequest.STATUS_CHOICES,
        'selected_status': status_filter,
        'query':           query,
        'user_role':       user_role,
    })


@login_required
def my_passport_requests(request):
    """Employee's own passport requests."""
    try:
        employee = request.user.employee
    except Exception:
        return render(request, 'hr_de/unauthorized.html', status=403)
    requests_qs = PassportRequest.objects.filter(employee=employee)
    return render(request, 'hr_de/passport/my_requests.html', {
        'employee': employee,
        'requests': requests_qs,
    })


#######################  AUTHENTICATION VIEWS #######################
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required

def signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm_password')
        email = request.POST.get('email')
        if password != confirm:
            messages.error(request, "Passwords do not match.")
            return redirect('signup')
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect('signup')
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken.")
            return redirect('signup')
        user = User.objects.create_user(username=username, password=password, first_name=first_name)
        login(request, user)
        return redirect('home')  # Change to your homepage view name
    return render(request, 'hr_de/auth/signup.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            user_role = getattr(user, 'role', None)
            user_employee = getattr(user, 'employee', None)

            if not user_role and not user_employee:
                messages.error(request, "You do not have permission to access the HR module.")
                return redirect('login')

            login(request, user)

            if user_role:
                if user_role.role == 'Head':
                    return redirect('employee_home')
                return redirect('home')
            # Regular employee (no Role) → self-service dashboard
            return redirect('employee_dashboard')
        else:
            messages.error(request, "Invalid credentials.")
    return render(request, 'hr_de/auth/login.html')

def logout_view(request):
    logout(request)
    return redirect('home')

from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
import random
import string

def generate_random_password():
    # Generate 4 random digits
    digits = ''.join(random.choices(string.digits, k=4))
    # Combine with 'junaid'
    return f'junaid{digits}'

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            new_password = generate_random_password()
            user.set_password(new_password)
            user.save()

            send_mail(
                subject='Your New Password',
                message=f'Username: {user.username} Your new login password is: {new_password}',
                from_email='noreply@example.com',
                recipient_list=[email],
                fail_silently=False,
            )
            messages.success(request, 'New password sent to your email.')
            return redirect('forgot_password')
        except User.DoesNotExist:
            messages.error(request, 'User with this email does not exist.')
    return render(request, 'hr_de/auth/forgot_password.html')

# ── DOCUMENT EXPIRY NOTIFICATION HELPERS ────────────────────────────────────

_DOC_FIELDS = [
    ('PASSPORT',        'passport_expiry',        'Passport',        'fas fa-passport'),
    ('VISA',            'visa_expiry',             'Visa',            'fas fa-stamp'),
    ('EID',             'eid_expiry',              'Emirates ID',     'fas fa-id-card'),
    ('LABOUR_CARD',     'labour_card_expiry',      'Labour Card',     'fas fa-briefcase'),
    ('INSURANCE',       'insurance_expiry',        'Insurance',       'fas fa-shield-halved'),
    ('DRIVING_LICENSE', 'driving_license_expiry',  'Driving License', 'fas fa-car'),
]


_MOL_DOC_FIELDS = [
    ('MOL_TRADE_LICENSE', 'trade_license_expiry',      'Trade License'),
    ('MOL_TENANCY',       'tenancy_contract_expiry',   'Tenancy Contract'),
    ('MOL_ESTABLISHMENT', 'establishment_card_expiry', 'Establishment Card'),
]


def _expiry_urgency_and_text(expiry, today, label, subject):
    days_left = (expiry - today).days
    if days_left < 0:
        urgency = 'critical'
        title   = f"{label} EXPIRED — {subject}"
        message = (
            f"{subject}'s {label} expired on {expiry.strftime('%d %b %Y')} "
            f"({abs(days_left)} day{'s' if abs(days_left) != 1 else ''} ago)."
        )
    elif days_left <= 7:
        urgency = 'critical'
        title   = f"{label} expiring in {days_left} day{'s' if days_left != 1 else ''} — {subject}"
        message = (
            f"{subject}'s {label} expires on {expiry.strftime('%d %b %Y')} "
            f"— only {days_left} day{'s' if days_left != 1 else ''} remaining."
        )
    elif days_left <= 30:
        urgency = 'warning'
        title   = f"{label} expiring soon — {subject}"
        message = (
            f"{subject}'s {label} expires on {expiry.strftime('%d %b %Y')} "
            f"({days_left} days left)."
        )
    else:
        urgency = 'info'
        title   = f"{label} expiry notice — {subject}"
        message = (
            f"{subject}'s {label} expires on {expiry.strftime('%d %b %Y')} "
            f"({days_left} days left)."
        )
    return days_left, urgency, title, message


_URGENCY_RANK = {'critical': 0, 'warning': 1, 'info': 2}


def _upsert_expiry_notification(*, category, doc_type, urgency, title, message,
                                employee=None, mol=None, vehicle=None, management_member=None):
    """Create/update ONE notification per document, dedup-safe across refreshes.

    Looks at the latest notification for this document regardless of read state:
      • none yet          → create it
      • exists, unread    → update text/urgency in place if it changed
      • exists, read      → suppress (do NOT recreate) unless the urgency has
                            escalated (e.g. warning → critical / expired), in
                            which case surface one fresh alert.
    """
    qs = Notification.objects.filter(category=category, doc_type=doc_type)
    qs = qs.filter(employee=employee) if employee is not None else qs.filter(employee__isnull=True)
    qs = qs.filter(mol=mol) if mol is not None else qs.filter(mol__isnull=True)
    qs = qs.filter(vehicle=vehicle) if vehicle is not None else qs.filter(vehicle__isnull=True)
    qs = qs.filter(management_member=management_member) if management_member is not None else qs.filter(management_member__isnull=True)
    existing = qs.order_by('-created_at').first()

    if existing is None:
        Notification.objects.create(
            employee=employee, mol=mol, vehicle=vehicle, management_member=management_member,
            title=title, message=message,
            category=category, urgency=urgency, doc_type=doc_type,
        )
        return

    if not existing.is_read:
        if existing.urgency != urgency or existing.title != title or existing.message != message:
            existing.urgency = urgency
            existing.title   = title
            existing.message = message
            existing.save(update_fields=['urgency', 'title', 'message'])
        return

    # Already read — only re-surface if it got more severe than what was seen.
    if _URGENCY_RANK.get(urgency, 3) < _URGENCY_RANK.get(existing.urgency, 3):
        Notification.objects.create(
            employee=employee, mol=mol, vehicle=vehicle, management_member=management_member,
            title=title, message=message,
            category=category, urgency=urgency, doc_type=doc_type,
        )


def _generate_expiry_notifications():
    today = timezone.now().date()
    # Keys of documents currently within the 60-day window; anything else that
    # still has an UNREAD alert has been renewed/cleared and must be pruned.
    active = set()

    # ── Employee document expiries ───────────────────────────────────────────
    for emp in Employee.objects.filter(is_active=True).select_related('department'):
        for doc_key, expiry_field, doc_label, _ in _DOC_FIELDS:
            expiry = getattr(emp, expiry_field, None)
            if not expiry:
                continue
            if (expiry - today).days > 60:
                continue
            _, urgency, title, message = _expiry_urgency_and_text(expiry, today, doc_label, emp.emp_name)
            _upsert_expiry_notification(
                category='document_expiry', doc_type=doc_key, employee=emp,
                urgency=urgency, title=title, message=message,
            )
            active.add(('document_expiry', doc_key, emp.id, None, None, None))

    # ── MOL document expiries ────────────────────────────────────────────────
    for mol in Mol.objects.all():
        for doc_key, expiry_field, doc_label in _MOL_DOC_FIELDS:
            expiry = getattr(mol, expiry_field, None)
            if not expiry:
                continue
            if (expiry - today).days > 60:
                continue
            _, urgency, title, message = _expiry_urgency_and_text(expiry, today, doc_label, mol.mol)
            _upsert_expiry_notification(
                category='mol_document_expiry', doc_type=doc_key, mol=mol,
                urgency=urgency, title=title, message=message,
            )
            active.add(('mol_document_expiry', doc_key, None, mol.id, None, None))

    # ── Vehicle Mulkiya expiries ─────────────────────────────────────────────
    for vehicle in Vehicle.objects.select_related('employee').all():
        expiry = vehicle.mulkiya_expiry
        if not expiry:
            continue
        if (expiry - today).days > 60:
            continue
        subject = f"{vehicle.name} ({vehicle.car_number})" if vehicle.car_number else vehicle.name
        _, urgency, title, message = _expiry_urgency_and_text(expiry, today, 'Mulkiya', subject)
        _upsert_expiry_notification(
            category='vehicle_document_expiry', doc_type='MULKIYA', vehicle=vehicle,
            urgency=urgency, title=title, message=message,
        )
        active.add(('vehicle_document_expiry', 'MULKIYA', None, None, vehicle.id, None))

    # ── Management member document expiries (EID / Passport / DL / Visa + country visas) ──
    _MGMT_DOCS = [
        ('MGMT_EID',      'eid_expiry',      'Emirates ID'),
        ('MGMT_PASSPORT', 'passport_expiry', 'Passport'),
        ('MGMT_DL',       'dl_expiry',       'Driving License'),
        ('MGMT_VISA',     'visa_expiry',     'Visa'),
    ]
    for m in ManagementMember.objects.select_related('head').prefetch_related('country_visas').all():
        subject = m.name if not m.is_family else f"{m.name} ({m.get_relation_display()} of {m.head.name})"
        for doc_key, field, label in _MGMT_DOCS:
            expiry = getattr(m, field, None)
            if not expiry or (expiry - today).days > 60:
                continue
            _, urgency, title, message = _expiry_urgency_and_text(expiry, today, label, subject)
            _upsert_expiry_notification(
                category='management_document_expiry', doc_type=doc_key, management_member=m,
                urgency=urgency, title=title, message=message,
            )
            active.add(('management_document_expiry', doc_key, None, None, None, m.id))
        for cv in m.country_visas.all():
            if not cv.expiry or (cv.expiry - today).days > 60:
                continue
            label = f"{cv.country} Visa"
            _, urgency, title, message = _expiry_urgency_and_text(cv.expiry, today, label, subject)
            doc_key = f'MGMT_COUNTRY_VISA_{cv.id}'
            _upsert_expiry_notification(
                category='management_document_expiry', doc_type=doc_key, management_member=m,
                urgency=urgency, title=title, message=message,
            )
            active.add(('management_document_expiry', doc_key, None, None, None, m.id))

    # ── Prune stale UNREAD alerts (document renewed, date cleared or removed) ──
    for n in (Notification.objects
              .filter(is_read=False,
                      category__in=['document_expiry', 'mol_document_expiry', 'vehicle_document_expiry', 'management_document_expiry'])
              .values('id', 'category', 'doc_type', 'employee_id', 'mol_id', 'vehicle_id', 'management_member_id')):
        key = (n['category'], n['doc_type'], n['employee_id'], n['mol_id'], n['vehicle_id'], n['management_member_id'])
        if key not in active:
            Notification.objects.filter(id=n['id']).delete()


def _refresh_vehicle_notifications(vehicle):
    """Keep a single vehicle's Mulkiya alert in sync right after it is edited.

    If the expiry was renewed (now >60 days away) or cleared, drop the stale
    UNREAD alert; otherwise upsert so the text reflects the new date."""
    # The caller's instance may hold raw POST strings; read canonical DB values.
    vehicle.refresh_from_db()
    today = timezone.now().date()
    stale = Notification.objects.filter(
        category='vehicle_document_expiry', doc_type='MULKIYA',
        vehicle=vehicle, is_read=False,
    )
    expiry = vehicle.mulkiya_expiry
    if not expiry or (expiry - today).days > 60:
        stale.delete()
        return
    subject = f"{vehicle.name} ({vehicle.car_number})" if vehicle.car_number else vehicle.name
    _, urgency, title, message = _expiry_urgency_and_text(expiry, today, 'Mulkiya', subject)
    _upsert_expiry_notification(
        category='vehicle_document_expiry', doc_type='MULKIYA', vehicle=vehicle,
        urgency=urgency, title=title, message=message,
    )


def _dedupe_read_notifications():
    """Remove duplicate READ expiry notifications left behind by the old logic —
    keep the newest read one per document, drop the rest."""
    seen, dupes = set(), []
    for n in (Notification.objects
              .filter(is_read=True,
                      category__in=['document_expiry', 'mol_document_expiry', 'vehicle_document_expiry', 'management_document_expiry'])
              .order_by('-created_at')
              .values('id', 'category', 'doc_type', 'employee_id', 'mol_id', 'vehicle_id', 'management_member_id')):
        key = (n['category'], n['doc_type'], n['employee_id'], n['mol_id'], n['vehicle_id'], n['management_member_id'])
        if key in seen:
            dupes.append(n['id'])
        else:
            seen.add(key)
    if dupes:
        Notification.objects.filter(id__in=dupes).delete()


def _cleanup_old_read_notifications():
    """Delete read notifications older than 30 days."""
    cutoff = timezone.now() - timezone.timedelta(days=30)
    Notification.objects.filter(is_read=True, created_at__lt=cutoff).delete()


@login_required
def notification_list(request):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    _generate_expiry_notifications()
    _dedupe_read_notifications()          # clean up any old duplicate read alerts
    _cleanup_old_read_notifications()     # drop read alerts older than 30 days

    _urgency_rank = {'critical': 0, 'warning': 1, 'info': 2}
    unread = sorted(
        Notification.objects.filter(is_read=False)
            .select_related('employee', 'employee__department', 'mol', 'vehicle', 'vehicle__employee', 'management_member', 'management_member__head'),
        key=lambda n: _urgency_rank.get(n.urgency, 3),
    )
    read = list(
        Notification.objects.filter(is_read=True)
            .select_related('employee', 'employee__department', 'mol', 'vehicle', 'vehicle__employee', 'management_member', 'management_member__head')
            .order_by('-created_at')[:50]
    )

    critical_count = sum(1 for n in unread if n.urgency == 'critical')
    warning_count  = sum(1 for n in unread if n.urgency == 'warning')

    return render(request, 'hr_de/notifications.html', {
        'unread':         unread,
        'read':           read,
        'unread_count':   len(unread),
        'critical_count': critical_count,
        'warning_count':  warning_count,
    })


@require_POST
def mark_notification_read(request, pk):
    try:
        n = Notification.objects.get(pk=pk)
        n.is_read = True
        n.save()
        _cleanup_old_read_notifications()
        return JsonResponse({'status': 'success'})
    except Notification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notification not found'}, status=404)


@require_POST
def mark_all_notifications_read(request):
    Notification.objects.filter(is_read=False).update(is_read=True)
    _cleanup_old_read_notifications()
    return JsonResponse({'status': 'success'})


############ BULK UPLOAD EMPLOYEE DATA ############

import pandas as pd
from datetime import datetime
from django.shortcuts import render, redirect
from .models import Employee, Mol
from django.contrib import messages

_BULK_COLUMNS = [
    ("EMP ID",                  "Required – unique employee code (e.g. EMP001)"),
    ("EMP NAME",                "Required – full name"),
    ("DOB",                     "Required – date of birth (YYYY-MM-DD)"),
    ("GENDER",                  "M or F"),
    ("NATIONALITY",             "Country code e.g. IN, AE, PH"),
    ("CONTACT NUMBER",          "Mobile / phone number"),
    ("DEPT",                    "Must match an existing Department name"),
    ("DESIGNATION",             "Job title"),
    ("MOL",                     "Ministry of Labour / sub-company name"),
    ("DATE OF JOIN",            "Joining date (YYYY-MM-DD)"),
    ("SALARY",                  "Monthly gross salary (number)"),
    # Salary structure components — if any are filled, the salary structure is set
    ("BASIC",                   "Salary component – basic pay (number)"),
    ("HRA",                     "Salary component – housing allowance (number)"),
    ("TRANSPORTATION",          "Salary component – transport allowance (number)"),
    ("OTHERS",                  "Salary component – other allowances (number)"),
    ("FUEL",                    "Salary component – fuel allowance (number)"),
    ("LOCATION",                "Work location / site"),
    ("EMPLOYEE STATUS",         "Active | Inactive | On Leave | Resigned | Terminated"),
    ("IS ACTIVE",               "YES or NO"),
    ("PASSPORT STATUS",         "With company | With employee"),
    ("PASSPORT NO.",            "Passport number"),
    ("PASSPORT EXPIRY",         "YYYY-MM-DD"),
    ("LABOUR CARD NO",          "Labour card number"),
    ("LABOUR EXPIRY",           "YYYY-MM-DD"),
    ("LABOUR NO",               "Labour No. — work permit / labour file no. (different from labour card no.)"),
    ("EID NO",                  "Emirates ID number"),
    ("EID EXPIRY",              "YYYY-MM-DD"),
    ("VISA NO",                 "Visa / residence permit number"),
    ("VISA EXPIRY",             "YYYY-MM-DD"),
    ("INSURANCE NO",            "Insurance policy number"),
    ("INSURANCE EXPIRY",        "YYYY-MM-DD"),
    ("DRIVING LICENSE NO",      "Driving licence number"),
    ("DRIVING LICENSE EXPIRY",  "YYYY-MM-DD"),
    # Bank details — for salary transfer (WPS)
    ("BANK NAME",               "Bank name (e.g. Emirates NBD)"),
    ("IBAN",                    "IBAN (e.g. AE07 0331 2345 6789 0123 456)"),
    ("ROUTING CODE",            "Bank routing / sort code"),
    # Login account — leave blank to skip account creation
    ("USERNAME",                "Optional – login username. Leave blank to skip account creation."),
    ("PASSWORD",                "Optional – initial password. Required if USERNAME is filled."),
]

# Column indices (0-based) that hold login credentials — styled differently in the template
_ACCOUNT_COL_NAMES = {"USERNAME", "PASSWORD"}

def _parse_date(val):
    if pd.isnull(val):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None

def download_employee_template(request):
    """Return a pre-formatted .xlsx template the user fills in and uploads."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_fill   = PatternFill("solid", fgColor="1E3A5F")
    acct_fill     = PatternFill("solid", fgColor="7C3AED")   # purple for account cols
    note_fill     = PatternFill("solid", fgColor="EBF3FF")
    acct_note_fill= PatternFill("solid", fgColor="F3E8FF")   # light purple note
    sample_fill   = PatternFill("solid", fgColor="F0FFF4")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    note_font     = Font(italic=True, color="1E3A5F", size=9)
    acct_note_font= Font(italic=True, color="6D28D9", size=9)
    sample_font   = Font(color="1E3A5F", size=10)
    thin_border   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [col for col, _ in _BULK_COLUMNS]
    notes   = [note for _, note in _BULK_COLUMNS]

    # Row 1 – column headers
    for col_idx, header in enumerate(headers, start=1):
        is_acct = header in _ACCOUNT_COL_NAMES
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = acct_fill if is_acct else header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = thin_border

    # Row 2 – notes / instructions
    for col_idx, (header, note) in enumerate(_BULK_COLUMNS, start=1):
        is_acct = header in _ACCOUNT_COL_NAMES
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.fill      = acct_note_fill if is_acct else note_fill
        cell.font      = acct_note_font if is_acct else note_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border    = thin_border

    # Row 3 – sample data
    sample = [
        "EMP001", "Ali Hassan", "1990-06-15", "M", "AE", "0501234567",
        "Sales", "Sales Manager", "ABC Trading LLC", "2023-01-01",
        "8000", "4000", "1500", "1000", "500", "1000",
        "Dubai", "Active", "YES", "With company",
        "P12345678", "2028-06-15", "LC123456", "2025-06-30", "L-2024-5567",
        "784-1990-1234567-1", "2027-03-15", "UAE-V-2024-778899", "2025-12-31",
        "INS001", "2026-06-01", "DL123456", "2027-01-01",
        "Emirates NBD", "AE070331234567890123456", "302620122",
        "ali.hassan", "Pass@1234",
    ]
    for col_idx, (val, (header, _)) in enumerate(zip(sample, _BULK_COLUMNS), start=1):
        is_acct = header in _ACCOUNT_COL_NAMES
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill      = PatternFill("solid", fgColor="EDE9FE") if is_acct else sample_fill
        cell.font      = Font(color="6D28D9", size=10) if is_acct else sample_font
        cell.border    = thin_border

    # Force account-like columns to Text so Excel keeps leading zeros the moment
    # a user types an IBAN / account / routing number that starts with 0.
    text_cols = {"IBAN", "ROUTING CODE"}
    text_idxs = [i for i, (h, _) in enumerate(_BULK_COLUMNS, start=1) if h in text_cols]
    for ci in text_idxs:
        for ri in range(3, 1001):          # sample row + plenty of data rows
            ws.cell(row=ri, column=ci).number_format = "@"

    # Row heights & column widths
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    col_widths = [12, 22, 14, 8, 14, 16, 18, 18, 22, 14,   # EMP ID .. DATE OF JOIN
                  10, 10, 10, 16, 10, 10,                   # SALARY, BASIC, HRA, TRANSPORTATION, OTHERS, FUEL
                  14, 16, 10, 18,                           # LOCATION, EMP STATUS, IS ACTIVE, PASSPORT STATUS
                  16, 16, 14, 14, 16, 22, 14, 20, 14,       # PASSPORT, LABOUR CARD NO, LABOUR EXPIRY, LABOUR NO, EID.. VISA
                  14, 16, 18, 20,                           # INSURANCE.. DRIVING LICENSE EXPIRY
                  18, 26, 14,                               # bank (name, IBAN, routing)
                  18, 16]                                   # login
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employee_upload_template.xlsx"'
    wb.save(response)
    return response


def bulk_upload_employees(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            # Read account-like columns as text so leading zeros survive (IBAN,
            # account & routing numbers often start with zeros).
            _text_cols = {"IBAN", "ROUTING CODE"}
            df = pd.read_excel(
                excel_file, header=0, skiprows=[1],   # skip the notes row
                dtype={c: str for c in _text_cols},
            )

            created_count    = 0
            updated_count    = 0
            accounts_created = 0
            error_rows       = []

            # New template splits labour card no. and labour no. into two columns.
            # Old templates only have "LABOUR NO" (which was the card number).
            has_labour_card_col = "LABOUR CARD NO" in df.columns

            for index, row in df.iterrows():
                try:
                    emp_id = str(row.get("EMP ID") or '').strip()
                    if not emp_id or emp_id.lower() == 'nan':
                        continue

                    mol_name = str(row.get("MOL") or '').strip()
                    mol_obj = None
                    if mol_name and mol_name.lower() != 'nan':
                        mol_obj, _ = Mol.objects.get_or_create(mol=mol_name)

                    dept_name = str(row.get("DEPT") or '').strip()
                    dept_obj = None
                    if dept_name and dept_name.lower() != 'nan':
                        dept_obj = Department.objects.filter(name__iexact=dept_name).first()

                    is_active_raw = str(row.get("IS ACTIVE") or 'YES').strip().upper()
                    is_active = is_active_raw not in ('NO', 'FALSE', '0', 'N')

                    gender_raw = str(row.get("GENDER") or 'M').strip().upper()
                    gender = gender_raw if gender_raw in ('M', 'F') else 'M'

                    status_raw = str(row.get("EMPLOYEE STATUS") or 'Active').strip()
                    valid_statuses = [c[0] for c in Employee.EMPLOYEE_STATUS_CHOICES]
                    employee_status = status_raw if status_raw in valid_statuses else 'Active'

                    def _str(key):
                        val = row.get(key)
                        if val is None or (isinstance(val, float) and pd.isnull(val)):
                            return None
                        return str(val).strip() or None

                    def _acct_str(key):
                        """Like _str but preserves leading zeros and drops a stray
                        trailing '.0' if Excel coerced the value to a number."""
                        val = row.get(key)
                        if val is None or (isinstance(val, float) and pd.isnull(val)):
                            return None
                        s = str(val).strip()
                        if s == '' or s.lower() == 'nan':
                            return None
                        if s.endswith('.0') and s[:-2].isdigit():
                            s = s[:-2]
                        return s

                    emp_obj, created = Employee.objects.update_or_create(
                        emp_id=emp_id,
                        defaults={
                            "emp_name":               str(row.get("EMP NAME") or '').strip(),
                            "dob":                    _parse_date(row.get("DOB")),
                            "gender":                 gender,
                            "nationality":            _str("NATIONALITY") or "IN",
                            "contact_number":         _str("CONTACT NUMBER"),
                            "department":             dept_obj,
                            "designation":            _str("DESIGNATION"),
                            "mol":                    mol_obj,
                            "joining_date":           _parse_date(row.get("DATE OF JOIN")),
                            "emp_salary":             row.get("SALARY") or None,
                            "job_location":           _str("LOCATION"),
                            "employee_status":        employee_status,
                            "is_active":              is_active,
                            "passport_status":        _str("PASSPORT STATUS") or "With company",
                            "passport_number":        _str("PASSPORT NO."),
                            "passport_expiry":        _parse_date(row.get("PASSPORT EXPIRY")),
                            "visa_number":            _str("VISA NO"),
                            "labour_card_number":     _str("LABOUR CARD NO") if has_labour_card_col else _str("LABOUR NO"),
                            "labour_number":          _str("LABOUR NO") if has_labour_card_col else None,
                            "labour_card_expiry":     _parse_date(row.get("LABOUR EXPIRY")),
                            "eid_number":             _str("EID NO"),
                            "eid_expiry":             _parse_date(row.get("EID EXPIRY")),
                            "visa_expiry":            _parse_date(row.get("VISA EXPIRY")),
                            "insurance_number":       _str("INSURANCE NO"),
                            "insurance_expiry":       _parse_date(row.get("INSURANCE EXPIRY")),
                            "driving_license_number": _str("DRIVING LICENSE NO"),
                            "driving_license_expiry": _parse_date(row.get("DRIVING LICENSE EXPIRY")),
                            "bank_name":              _str("BANK NAME"),
                            "iban":                   _acct_str("IBAN"),
                            "routing_code":           _acct_str("ROUTING CODE"),
                        }
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                    # ── Salary structure (set if any component is filled) ─────
                    def _num(key):
                        val = row.get(key)
                        if val is None or (isinstance(val, float) and pd.isnull(val)):
                            return 0.0
                        try:
                            return float(str(val).strip())
                        except (ValueError, TypeError):
                            return 0.0

                    s_basic = _num("BASIC")
                    s_hra   = _num("HRA")
                    s_trans = _num("TRANSPORTATION")
                    s_other = _num("OTHERS")
                    s_fuel  = _num("FUEL")
                    if any([s_basic, s_hra, s_trans, s_other, s_fuel]):
                        SalaryStructure.objects.update_or_create(
                            employee=emp_obj,
                            defaults={
                                "basic": s_basic, "hra": s_hra,
                                "transport": s_trans, "others": s_other,
                                "fuel": s_fuel, "updated_by": request.user,
                            },
                        )

                    # ── Login account creation ───────────────────────────────
                    username = _str("USERNAME")
                    password = _str("PASSWORD")
                    if username and password:
                        if emp_obj.user_id:
                            # Employee already has an account — leave it alone
                            pass
                        elif User.objects.filter(username=username).exists():
                            error_rows.append(
                                f"Row {index + 3} ({emp_id}): username '{username}' is already taken — account not created."
                            )
                        else:
                            new_user = User.objects.create_user(
                                username=username,
                                password=password,
                                first_name=emp_obj.emp_name,
                            )
                            emp_obj.user = new_user
                            emp_obj.save(update_fields=["user"])
                            accounts_created += 1

                except Exception as e:
                    error_rows.append(f"Row {index + 3}: {e}")

            if error_rows:
                for err in error_rows:
                    messages.error(request, err)

            if created_count or updated_count:
                acct_note = f", {accounts_created} account{'s' if accounts_created != 1 else ''} created" if accounts_created else ""
                messages.success(
                    request,
                    f"Upload complete — {created_count} created, {updated_count} updated{acct_note}."
                )
            return redirect("employee_upload")

        except Exception as e:
            messages.error(request, f"Upload failed: {e}")

    columns = [col for col, _ in _BULK_COLUMNS]
    return render(request, "hr_de/employee_bulk_upload.html", {"columns": columns})


# ── MD System Account Management (HR only) ──────────────────────────────────

@login_required
def md_account_list(request):
    try:
        caller_role = request.user.role.role
    except Exception:
        caller_role = None
    if caller_role != 'HR':
        return render(request, 'hr_de/unauthorized.html', status=403)

    md_roles = Role.objects.filter(role='MD').select_related('user')
    return render(request, 'hr_de/md_accounts.html', {'md_roles': md_roles})


@login_required
def md_account_create(request):
    try:
        caller_role = request.user.role.role
    except Exception:
        caller_role = None
    if caller_role != 'HR':
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        full_name = request.POST.get('full_name', '').strip()

        if not username or not password:
            messages.error(request, "Username and password are required.")
            return redirect('md_account_list')

        if User.objects.filter(username=username).exists():
            messages.error(request, f"Username '{username}' is already taken.")
            return redirect('md_account_list')

        new_user = User.objects.create_user(
            username=username,
            password=password,
            first_name=full_name,
        )
        Role.objects.create(user=new_user, role='MD')
        messages.success(request, f"MD account '{username}' created successfully.")
        return redirect('md_account_list')

    return redirect('md_account_list')


@login_required
def md_account_delete(request, pk):
    try:
        caller_role = request.user.role.role
    except Exception:
        caller_role = None
    if caller_role != 'HR':
        return render(request, 'hr_de/unauthorized.html', status=403)

    role_obj = get_object_or_404(Role, pk=pk, role='MD')
    user_obj = role_obj.user
    username = user_obj.username if user_obj else '—'
    user_obj.delete()
    messages.success(request, f"MD account '{username}' deleted.")
    return redirect('md_account_list')


# ── SALARY REVISION (AUDIT TRAIL) ────────────────────────────────────────────

@login_required
def salary_revision_create(request, emp_id):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    employee  = get_object_or_404(Employee, pk=emp_id)
    try:
        structure = employee.salary_structure
    except SalaryStructure.DoesNotExist:
        structure = None

    if request.method == 'POST':
        def _f(key):
            try:
                return round(float(request.POST.get(key, 0) or 0), 2)
            except (ValueError, TypeError):
                return 0.0

        new_basic     = _f('basic')
        new_hra       = _f('hra')
        new_transport = _f('transport')
        new_fuel      = _f('fuel')
        new_others    = _f('others')
        new_salary    = round(new_basic + new_hra + new_transport + new_fuel + new_others, 2)

        # Effective date is captured as a month (YYYY-MM); store the 1st of it.
        effective_month = (request.POST.get('effective_date') or '').strip()
        try:
            effective_date = datetime.strptime(effective_month + '-01', '%Y-%m-%d').date()
        except ValueError:
            try:
                # Fall back to a full date if one was somehow submitted.
                effective_date = datetime.strptime(effective_month, '%Y-%m-%d').date()
            except ValueError:
                effective_date = timezone.now().date().replace(day=1)
        reason         = request.POST.get('reason', '').strip()

        old_salary    = employee.emp_salary or 0
        old_basic     = structure.basic     if structure else 0
        old_hra       = structure.hra       if structure else 0
        old_transport = structure.transport if structure else 0
        old_fuel      = structure.fuel      if structure else 0
        old_others    = structure.others    if structure else 0

        if new_salary > old_salary:
            change_type = 'Increment'
        elif new_salary < old_salary:
            change_type = 'Decrement'
        else:
            change_type = 'Adjustment'

        SalaryRevision.objects.create(
            employee=employee,
            change_type=change_type,
            effective_date=effective_date,
            old_salary=old_salary,
            old_basic=old_basic,
            old_hra=old_hra,
            old_transport=old_transport,
            old_fuel=old_fuel,
            old_others=old_others,
            new_salary=new_salary,
            new_basic=new_basic,
            new_hra=new_hra,
            new_transport=new_transport,
            new_fuel=new_fuel,
            new_others=new_others,
            reason=reason,
            changed_by=request.user,
        )

        employee.emp_salary = new_salary
        employee.save(update_fields=['emp_salary', 'updated_at'])

        if structure:
            structure.basic      = new_basic
            structure.hra        = new_hra
            structure.transport  = new_transport
            structure.fuel       = new_fuel
            structure.others     = new_others
            structure.updated_by = request.user
            structure.save()
        else:
            SalaryStructure.objects.create(
                employee=employee, basic=new_basic, hra=new_hra,
                transport=new_transport, fuel=new_fuel, others=new_others,
                updated_by=request.user,
            )

        messages.success(request, f"Salary {change_type.lower()} applied for {employee.emp_name}.")
        return redirect('salary_history', emp_id=emp_id)

    return render(request, 'hr_de/salary/revision_form.html', {
        'employee':  employee,
        'structure': structure,
        'today':     timezone.now().date().isoformat(),
        'today_month': timezone.now().strftime('%Y-%m'),
    })


@login_required
def salary_history(request, emp_id):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    employee  = get_object_or_404(Employee, pk=emp_id)
    revisions = employee.salary_revisions.select_related('changed_by').order_by('-effective_date', '-changed_at')

    return render(request, 'hr_de/salary/history.html', {
        'employee':  employee,
        'revisions': revisions,
    })


@login_required
def salary_all_revisions(request):
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    qs = SalaryRevision.objects.select_related(
        'employee', 'employee__department', 'changed_by'
    ).order_by('-effective_date', '-changed_at')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(employee__emp_name__icontains=q) | Q(employee__emp_id__icontains=q)
        )

    return render(request, 'hr_de/salary/all_revisions.html', {
        'revisions': qs,
        'q': q,
    })


# ── ATTENDANCE ───────────────────────────────────────────────────────────────

_ATT_STATUSES = {'present', 'absent', 'half_day', 'leave', 'holiday', 'week_off'}


def _employee_for_user(user):
    """Return the Employee linked to this login, or None."""
    try:
        return user.employee
    except Employee.DoesNotExist:
        return None


@login_required
def attendance_mark(request):
    """Employee self-service: one-tap check-in stamping the current time — TODAY only."""
    employee = _employee_for_user(request.user)
    today    = timezone.localdate()

    if request.method == 'POST':
        if employee is None:
            messages.error(request, "Your login is not linked to an employee profile.")
            return redirect('attendance_mark')
        existing = Attendance.objects.filter(employee=employee, date=today).first()
        if existing:
            messages.info(request, "You've already checked in for today.")
        else:
            now_time = timezone.localtime().time()
            Attendance.objects.create(
                employee=employee, date=today, status='present',
                check_in=now_time, source='self', marked_by=request.user,
            )
            messages.success(request, f"Checked in at {now_time.strftime('%I:%M %p')}. Have a great day!")
        return redirect('attendance_mark')

    todays = None
    recent = []
    if employee is not None:
        todays = Attendance.objects.filter(employee=employee, date=today).first()
        recent = list(
            Attendance.objects.filter(employee=employee)
            .select_related('leave_type')
            .order_by('-date')[:14]
        )

    return render(request, 'hr_de/attendance/mark.html', {
        'employee': employee,
        'today':    today,
        'todays':   todays,
        'recent':   recent,
        'now':      timezone.localtime(),
    })


@login_required
def attendance_team(request):
    """Department head marks TODAY's attendance for the employees under them.

    Meant for support-staff teams whose members don't self check-in. Heads see
    only their own department and can only set today's attendance."""
    dept = _head_department(request)
    if dept is None:
        return render(request, 'hr_de/unauthorized.html', status=403)

    today = timezone.localdate()
    employees = list(
        Employee.objects.filter(department=dept, is_active=True).order_by('emp_name')
    )
    emp_ids = [e.id for e in employees]

    # Today's explicit attendance records
    att_map = {
        a.employee_id: a
        for a in Attendance.objects.filter(employee_id__in=emp_ids, date=today)
        .select_related('leave_type')
    }

    # Approved-leave overlay for today (read-only — shown but not editable)
    leave_map = {}
    for lv in Leave.objects.filter(
        employee_id__in=emp_ids, status='Approved'
    ).select_related('leave_type'):
        f = lv.actual_from or lv.expected_from
        t = lv.actual_to or lv.expected_to
        if f and t and f <= today <= t:
            leave_map[lv.employee_id] = lv.leave_type.name if lv.leave_type else 'Leave'

    rows = []
    for e in employees:
        a = att_map.get(e.id)
        if a:
            status = a.status
            leave_type = a.leave_type.name if a.leave_type else ''
            check_in = a.check_in
            auto_leave = ''
        elif e.id in leave_map:
            status, leave_type, check_in, auto_leave = 'leave', leave_map[e.id], None, leave_map[e.id]
        else:
            status, leave_type, check_in, auto_leave = '', '', None, ''
        rows.append({
            'employee':   e,
            'status':     status,
            'leave_type': leave_type,
            'check_in':   check_in,
            'auto_leave': auto_leave,
        })

    return render(request, 'hr_de/attendance/team.html', {
        'department':  dept,
        'today':       today,
        'rows':        rows,
        'leave_types': list(LeaveType.objects.all().order_by('name')),
        'now':         timezone.localtime(),
    })


@require_POST
@login_required
def attendance_team_mark(request):
    """AJAX: department head upserts TODAY's attendance for one of their employees."""
    dept = _head_department(request)
    if dept is None:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    import json
    try:
        payload = json.loads(request.body.decode('utf-8'))
        emp_id          = int(payload['employee_id'])
        status          = (payload.get('status') or '').strip()
        leave_type_name = (payload.get('leave_type') or '').strip()
    except (KeyError, ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid request data.'}, status=400)

    if status and status not in _ATT_STATUSES:
        return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)

    # Head may only touch employees in their own department
    employee = Employee.objects.filter(pk=emp_id, department=dept).first()
    if employee is None:
        return JsonResponse({'status': 'error', 'message': 'Employee not in your team.'}, status=403)

    today = timezone.localdate()

    # Empty status clears the cell
    if not status:
        Attendance.objects.filter(employee=employee, date=today).delete()
        return JsonResponse({'status': 'success', 'cleared': True})

    leave_type = None
    if status == 'leave':
        if not leave_type_name:
            return JsonResponse({'status': 'error', 'message': 'Pick a leave type.'}, status=400)
        leave_type = LeaveType.objects.filter(name=leave_type_name).first()
        if leave_type is None:
            return JsonResponse({'status': 'error', 'message': 'Unknown leave type.'}, status=400)

    check_in = timezone.localtime().time() if status == 'present' else None
    att, created = Attendance.objects.update_or_create(
        employee=employee, date=today,
        defaults={
            'status':     status,
            'leave_type': leave_type,
            'check_in':   check_in,
            'source':     'head',
            'marked_by':  request.user,
        },
    )
    return JsonResponse({
        'status':      'success',
        'created':     created,
        'cell_status': att.status,
        'leave_type':  leave_type.name if leave_type else '',
    })


@login_required
def attendance_grid(request):
    """HR month grid — department-wise, editable, with an approved-leave overlay."""
    if not _hr_only(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    import calendar
    import datetime as dt
    from collections import OrderedDict

    today = timezone.localdate()
    try:
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        year = int(request.GET.get('year', today.year))
    except (ValueError, TypeError):
        year = today.year
    if not (1 <= month <= 12):
        month = today.month

    num_days  = calendar.monthrange(year, month)[1]
    first_day = dt.date(year, month, 1)
    last_day  = dt.date(year, month, num_days)

    # Day-header meta (Sunday → weekly off)
    day_meta = []
    for d in range(1, num_days + 1):
        dobj = dt.date(year, month, d)
        day_meta.append({
            'day':       d,
            'weekday':   dobj.strftime('%a'),
            'is_off':    dobj.weekday() == 6,
            'is_future': dobj > today,
            'is_today':  dobj == today,
        })

    # Employees (optionally filtered by department)
    dept_id   = request.GET.get('department', '').strip()
    employees = (
        Employee.objects.filter(is_active=True)
        .select_related('department')
        .order_by('department__name', 'emp_name')
    )
    if dept_id:
        employees = employees.filter(department_id=dept_id)
    employees = list(employees)
    emp_ids   = [e.id for e in employees]

    # Explicit attendance records for the month
    att_map = {}
    for a in Attendance.objects.filter(
        employee_id__in=emp_ids, date__range=(first_day, last_day)
    ).select_related('leave_type'):
        att_map[(a.employee_id, a.date.day)] = a

    # Approved-leave overlay (fills only cells without an explicit record)
    leave_map = {}
    for lv in Leave.objects.filter(
        employee_id__in=emp_ids, status='Approved'
    ).select_related('leave_type'):
        f = lv.actual_from or lv.expected_from
        t = lv.actual_to or lv.expected_to
        if not f or not t:
            continue
        cur = max(f, first_day)
        end = min(t, last_day)
        while cur <= end:
            leave_map.setdefault(
                (lv.employee_id, cur.day),
                lv.leave_type.name if lv.leave_type else 'Leave',
            )
            cur += dt.timedelta(days=1)

    # Build rows grouped by department
    grouped = OrderedDict()
    for emp in employees:
        cells = []
        present_days = leave_days = absent_days = 0
        for dm in day_meta:
            d = dm['day']
            a = att_map.get((emp.id, d))
            if a:
                status, source, check_in = a.status, a.source, a.check_in
                lt = a.leave_type.name if a.leave_type else ''
            elif (emp.id, d) in leave_map:
                status, source, check_in = 'leave', 'leave_auto', None
                lt = leave_map[(emp.id, d)]
            elif dm['is_off']:
                status, source, check_in, lt = 'week_off', 'auto', None, ''
            else:
                status, source, check_in, lt = '', '', None, ''

            if   status == 'present': present_days += 1
            elif status == 'leave':   leave_days   += 1
            elif status == 'absent':  absent_days  += 1

            cells.append({
                'day':        d,
                'status':     status,
                'leave_type': lt,
                'source':     source,
                'is_off':     dm['is_off'],
                'is_future':  dm['is_future'],
                'is_today':   dm['is_today'],
                'check_in':   check_in,
            })

        dept_name = emp.department.name if emp.department else 'Unassigned'
        grouped.setdefault(dept_name, []).append({
            'employee':     emp,
            'cells':        cells,
            'present_days': present_days,
            'leave_days':   leave_days,
            'absent_days':  absent_days,
        })

    dept_groups = [{'name': name, 'rows': rows} for name, rows in grouped.items()]

    months = [{'num': i, 'name': dt.date(2000, i, 1).strftime('%B')} for i in range(1, 13)]
    years  = list(range(today.year - 3, today.year + 2))

    return render(request, 'hr_de/attendance/grid.html', {
        'dept_groups':   dept_groups,
        'day_meta':      day_meta,
        'num_days':      num_days,
        'month':         month,
        'year':          year,
        'months':        months,
        'years':         years,
        'leave_types':   list(LeaveType.objects.all().order_by('name')),
        'departments':   Department.objects.all().order_by('name'),
        'selected_dept': dept_id,
        'today':         today,
    })


@require_POST
def attendance_cell_update(request):
    """AJAX upsert for a single grid cell (HR only)."""
    if not _hr_only(request):
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    import json
    import datetime as dt

    try:
        payload = json.loads(request.body.decode('utf-8'))
        emp_id          = int(payload['employee_id'])
        year            = int(payload['year'])
        month           = int(payload['month'])
        day             = int(payload['day'])
        status          = (payload.get('status') or '').strip()
        leave_type_name = (payload.get('leave_type') or '').strip()
    except (KeyError, ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid request data.'}, status=400)

    if status and status not in _ATT_STATUSES:
        return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)

    try:
        date = dt.date(year, month, day)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Invalid date.'}, status=400)

    employee = get_object_or_404(Employee, pk=emp_id)

    # Empty status clears the cell (reverts to auto / blank)
    if not status:
        Attendance.objects.filter(employee=employee, date=date).delete()
        return JsonResponse({'status': 'success', 'cleared': True})

    leave_type = None
    if status == 'leave':
        if not leave_type_name:
            return JsonResponse({'status': 'error', 'message': 'Pick a leave type.'}, status=400)
        leave_type = LeaveType.objects.filter(name=leave_type_name).first()
        if leave_type is None:
            return JsonResponse({'status': 'error', 'message': 'Unknown leave type.'}, status=400)

    att, created = Attendance.objects.update_or_create(
        employee=employee, date=date,
        defaults={
            'status':     status,
            'leave_type': leave_type,
            'source':     'hr',
            'marked_by':  request.user,
        },
    )
    return JsonResponse({
        'status':      'success',
        'created':     created,
        'cell_status': att.status,
        'leave_type':  leave_type.name if leave_type else '',
    })


@require_POST
def attendance_bulk_update(request):
    """AJAX fill: apply one status to many days of a single employee (HR only)."""
    if not _hr_only(request):
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    import json
    import datetime as dt

    try:
        payload = json.loads(request.body.decode('utf-8'))
        emp_id          = int(payload['employee_id'])
        year            = int(payload['year'])
        month           = int(payload['month'])
        days            = [int(d) for d in payload['days']]
        status          = (payload.get('status') or '').strip()
        leave_type_name = (payload.get('leave_type') or '').strip()
    except (KeyError, ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid request data.'}, status=400)

    if status and status not in _ATT_STATUSES:
        return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)

    employee = get_object_or_404(Employee, pk=emp_id)

    leave_type = None
    if status == 'leave':
        if not leave_type_name:
            return JsonResponse({'status': 'error', 'message': 'Pick a leave type.'}, status=400)
        leave_type = LeaveType.objects.filter(name=leave_type_name).first()
        if leave_type is None:
            return JsonResponse({'status': 'error', 'message': 'Unknown leave type.'}, status=400)

    applied = []
    for d in days:
        try:
            date = dt.date(year, month, d)
        except ValueError:
            continue
        if not status:
            Attendance.objects.filter(employee=employee, date=date).delete()
        else:
            Attendance.objects.update_or_create(
                employee=employee, date=date,
                defaults={
                    'status':     status,
                    'leave_type': leave_type,
                    'source':     'hr',
                    'marked_by':  request.user,
                },
            )
        applied.append(d)

    return JsonResponse({'status': 'success', 'applied': applied})


@require_POST
def attendance_day_update(request):
    """AJAX: apply one status to many employees on a single date (e.g. a public holiday)."""
    if not _hr_only(request):
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    import json
    import datetime as dt

    try:
        payload = json.loads(request.body.decode('utf-8'))
        emp_ids         = [int(x) for x in payload['employee_ids']]
        year            = int(payload['year'])
        month           = int(payload['month'])
        day             = int(payload['day'])
        status          = (payload.get('status') or '').strip()
        leave_type_name = (payload.get('leave_type') or '').strip()
    except (KeyError, ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid request data.'}, status=400)

    if status and status not in _ATT_STATUSES:
        return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)

    try:
        date = dt.date(year, month, day)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Invalid date.'}, status=400)

    leave_type = None
    if status == 'leave':
        if not leave_type_name:
            return JsonResponse({'status': 'error', 'message': 'Pick a leave type.'}, status=400)
        leave_type = LeaveType.objects.filter(name=leave_type_name).first()
        if leave_type is None:
            return JsonResponse({'status': 'error', 'message': 'Unknown leave type.'}, status=400)

    applied = 0
    for emp in Employee.objects.filter(id__in=emp_ids):
        if not status:
            Attendance.objects.filter(employee=emp, date=date).delete()
        else:
            Attendance.objects.update_or_create(
                employee=emp, date=date,
                defaults={
                    'status':     status,
                    'leave_type': leave_type,
                    'source':     'hr',
                    'marked_by':  request.user,
                },
            )
        applied += 1

    return JsonResponse({'status': 'success', 'applied': applied})


# ── Other Records (document store — HR & MD only) ────────────────────────────

@login_required
def other_records_list(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    records = (OtherRecord.objects
               .prefetch_related('employees')
               .select_related('uploaded_by')
               .all())

    query = (request.GET.get('q') or '').strip()
    if query:
        records = records.filter(
            Q(title__icontains=query) |
            Q(comment__icontains=query) |
            Q(expense_other__icontains=query) |
            Q(employees__emp_name__icontains=query)
        ).distinct()

    expense = (request.GET.get('expense') or '').strip()
    if expense in dict(OtherRecord.EXPENSE_CHOICES):
        records = records.filter(expense_category=expense)

    return render(request, 'hr_de/other_records/list.html', {
        'records': records,
        'query':   query,
        'expense': expense,
        'expense_choices': OtherRecord.EXPENSE_CHOICES,
    })


@login_required
def other_record_add(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        if not title:
            messages.error(request, 'A heading is required.')
            return redirect('other_record_add')

        category = (request.POST.get('expense_category') or '').strip()
        if category not in dict(OtherRecord.EXPENSE_CHOICES):
            category = ''
        expense_other = (request.POST.get('expense_other') or '').strip()
        if category == 'other' and not expense_other:
            messages.error(request, 'Please specify the "Other" expense.')
            return redirect('other_record_add')

        rec = OtherRecord.objects.create(
            title=title,
            comment=(request.POST.get('comment') or '').strip(),
            expense_category=category,
            expense_other=expense_other if category == 'other' else '',
            document=request.FILES.get('document'),
            uploaded_by=request.user,
        )
        emp_ids = request.POST.getlist('employees')
        if emp_ids:
            rec.employees.set(Employee.objects.filter(pk__in=emp_ids))
        messages.success(request, f"Record '{rec.title}' saved.")
        return redirect('other_records_list')

    return render(request, 'hr_de/other_records/form.html', {
        'employees': Employee.objects.filter(is_active=True).order_by('emp_name'),
        'expense_choices': OtherRecord.EXPENSE_CHOICES,
    })


@login_required
def other_record_edit(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    rec = get_object_or_404(OtherRecord, pk=pk)

    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        if not title:
            messages.error(request, 'A heading is required.')
            return redirect('other_record_edit', pk=pk)

        category = (request.POST.get('expense_category') or '').strip()
        if category not in dict(OtherRecord.EXPENSE_CHOICES):
            category = ''
        expense_other = (request.POST.get('expense_other') or '').strip()
        if category == 'other' and not expense_other:
            messages.error(request, 'Please specify the "Other" expense.')
            return redirect('other_record_edit', pk=pk)

        rec.title = title
        rec.comment = (request.POST.get('comment') or '').strip()
        rec.expense_category = category
        rec.expense_other = expense_other if category == 'other' else ''
        if request.FILES.get('document'):
            rec.document = request.FILES['document']
        rec.save()
        rec.employees.set(Employee.objects.filter(pk__in=request.POST.getlist('employees')))
        messages.success(request, f"Record '{rec.title}' updated.")
        return redirect('other_records_list')

    return render(request, 'hr_de/other_records/form.html', {
        'record':    rec,
        'employees': Employee.objects.filter(is_active=True).order_by('emp_name'),
        'linked_ids': set(rec.employees.values_list('pk', flat=True)),
        'expense_choices': OtherRecord.EXPENSE_CHOICES,
    })


@require_POST
@login_required
def other_record_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    rec = get_object_or_404(OtherRecord, pk=pk)
    title = rec.title
    rec.delete()
    messages.success(request, f"Record '{title}' deleted.")
    return redirect('other_records_list')


# ── Vehicle Management (fleet register — HR & MD only) ───────────────────────

_VEHICLE_TEXT_FIELDS = [
    'name', 'car_number', 'model', 'tracking', 'state',
    'traffic_code', 'mortgage', 'car_and_model', 'company',
]


def _apply_vehicle_post(vehicle, request):
    """Copy submitted fields onto a Vehicle instance (shared by add & edit)."""
    for f in _VEHICLE_TEXT_FIELDS:
        setattr(vehicle, f, (request.POST.get(f) or '').strip())

    ownership = request.POST.get('ownership')
    if ownership in dict(Vehicle.OWNERSHIP_CHOICES):
        vehicle.ownership = ownership

    vehicle.tracking_exp_date = request.POST.get('tracking_exp_date') or None
    vehicle.mulkiya_expiry    = request.POST.get('mulkiya_expiry') or None

    emp_id = request.POST.get('employee')
    vehicle.employee = Employee.objects.filter(pk=emp_id).first() if emp_id else None

    if request.FILES.get('mulkiya_document'):
        vehicle.mulkiya_document = request.FILES['mulkiya_document']

    # Sale / disposal details — only kept when the vehicle is marked as sold.
    vehicle.is_sold = request.POST.get('is_sold') in ('1', 'on', 'true', 'True')
    if vehicle.is_sold:
        vehicle.sold_on     = request.POST.get('sold_on') or None
        vehicle.sold_amount = (request.POST.get('sold_amount') or '').strip() or None
        vehicle.sold_to     = (request.POST.get('sold_to') or '').strip()
        if request.FILES.get('sold_document'):
            vehicle.sold_document = request.FILES['sold_document']
    else:
        vehicle.sold_on     = None
        vehicle.sold_amount = None
        vehicle.sold_to     = ''
        # Note: the uploaded sale document is retained on file even if un-marked.


@login_required
def vehicle_list(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    from django.db.models import Count
    vehicles = Vehicle.objects.select_related('created_by', 'employee').annotate(
        pending_services=Count('services', filter=Q(services__status='Requested'))
    ).all()

    ownership = (request.GET.get('ownership') or '').strip()
    if ownership in dict(Vehicle.OWNERSHIP_CHOICES):
        vehicles = vehicles.filter(ownership=ownership)

    # Distinct, non-empty company names for the filter dropdown.
    companies = list(
        Vehicle.objects.exclude(company='')
        .values_list('company', flat=True)
        .order_by('company')
        .distinct()
    )
    company = (request.GET.get('company') or '').strip()
    if company:
        vehicles = vehicles.filter(company=company)

    # Distinct, non-empty states for the filter dropdown.
    states = list(
        Vehicle.objects.exclude(state='')
        .values_list('state', flat=True)
        .order_by('state')
        .distinct()
    )
    state = (request.GET.get('state') or '').strip()
    if state:
        vehicles = vehicles.filter(state=state)

    query = (request.GET.get('q') or '').strip()
    if query:
        vehicles = vehicles.filter(
            Q(name__icontains=query) |
            Q(car_number__icontains=query) |
            Q(model__icontains=query) |
            Q(car_and_model__icontains=query) |
            Q(company__icontains=query) |
            Q(traffic_code__icontains=query) |
            Q(employee__emp_name__icontains=query) |
            Q(employee__emp_id__icontains=query)
        ).distinct()

    # Sorting — map the sort key sent by the table headers to a real ORM field.
    SORT_FIELDS = {
        'name':              'name',
        'ownership':         'ownership',
        'employee':          'employee__emp_name',
        'car_number':        'car_number',
        'model':             'model',
        'car_and_model':     'car_and_model',
        'company':           'company',
        'tracking':          'tracking',
        'tracking_exp_date': 'tracking_exp_date',
        'state':             'state',
        'traffic_code':      'traffic_code',
        'mortgage':          'mortgage',
        'mulkiya_expiry':    'mulkiya_expiry',
    }
    sort = (request.GET.get('sort') or '').strip()
    direction = 'desc' if (request.GET.get('dir') or '').strip() == 'desc' else 'asc'
    if sort in SORT_FIELDS:
        order_field = SORT_FIELDS[sort]
        if direction == 'desc':
            order_field = '-' + order_field
        vehicles = vehicles.order_by(order_field)

    return render(request, 'hr_de/vehicles/list.html', {
        'vehicles':        vehicles,
        'query':           query,
        'ownership':       ownership,
        'companies':       companies,
        'company':         company,
        'states':          states,
        'state':           state,
        'sort':            sort,
        'dir':             direction,
        'company_count':   Vehicle.objects.filter(ownership='Company').count(),
        'personal_count':  Vehicle.objects.filter(ownership='Personal').count(),
    })


@login_required
def vehicle_add(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        car_number = (request.POST.get('car_number') or '').strip()
        if not name or not car_number:
            messages.error(request, 'Name and Car Number are required.')
            return redirect('vehicle_add')
        vehicle = Vehicle(created_by=request.user)
        _apply_vehicle_post(vehicle, request)
        vehicle.save()
        messages.success(request, f"Vehicle '{vehicle.name}' added.")
        return redirect('vehicle_list')

    return render(request, 'hr_de/vehicles/form.html', {
        'ownership_choices': Vehicle.OWNERSHIP_CHOICES,
        'employees': Employee.objects.filter(is_active=True).order_by('emp_name'),
    })


@login_required
def vehicle_edit(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    vehicle = get_object_or_404(Vehicle, pk=pk)

    # Where to return after saving/cancelling — e.g. opened from Notifications.
    next_target = request.POST.get('next') or request.GET.get('next') or ''
    back_url = 'notification_list' if next_target == 'notifications' else 'vehicle_list'

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        car_number = (request.POST.get('car_number') or '').strip()
        if not name or not car_number:
            messages.error(request, 'Name and Car Number are required.')
            return redirect(request.get_full_path())
        _apply_vehicle_post(vehicle, request)
        vehicle.save()
        # An expiry change may make the old alert stale — refresh notifications.
        _refresh_vehicle_notifications(vehicle)
        messages.success(request, f"Vehicle '{vehicle.name}' updated.")
        return redirect(back_url)

    return render(request, 'hr_de/vehicles/form.html', {
        'vehicle':           vehicle,
        'ownership_choices': Vehicle.OWNERSHIP_CHOICES,
        'employees': Employee.objects.filter(is_active=True).order_by('emp_name'),
        'next':              next_target,
        'back_url':          back_url,
    })


@login_required
def vehicle_detail(request, pk):
    """Read-only detail view for a single vehicle (HR/MD only)."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    vehicle = get_object_or_404(
        Vehicle.objects.select_related('employee', 'created_by'), pk=pk
    )
    services = vehicle.services.select_related('requested_by', 'approved_by')
    return render(request, 'hr_de/vehicles/detail.html', {
        'vehicle':      vehicle,
        'services':     services,
        'last_dates':   vehicle.last_service_dates(),
        'type_choices': VehicleService.SERVICE_TYPE_CHOICES,
        'today':        timezone.localdate(),
    })


@require_POST
@login_required
def vehicle_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    vehicle = get_object_or_404(Vehicle, pk=pk)
    name = vehicle.name
    vehicle.delete()
    messages.success(request, f"Vehicle '{name}' deleted.")
    return redirect('vehicle_list')


# ── Vehicle Services (maintenance history + employee requests) ───────────────

def _valid_service_type(value):
    return value in dict(VehicleService.SERVICE_TYPE_CHOICES)


@login_required
def vehicle_services(request, pk):
    """HR/MD: full service history for one vehicle + pending requests + log form."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    vehicle = get_object_or_404(Vehicle.objects.select_related('employee'), pk=pk)

    services = vehicle.services.select_related('requested_by', 'created_by', 'approved_by')
    pending  = services.filter(status='Requested')
    history  = services.exclude(status='Requested')

    return render(request, 'hr_de/vehicles/services.html', {
        'vehicle':       vehicle,
        'pending':       pending,
        'history':       history,
        'type_choices':  VehicleService.SERVICE_TYPE_CHOICES,
        'status_choices': VehicleService.STATUS_CHOICES,
        'last_dates':    vehicle.last_service_dates(),
        'today':         timezone.localdate(),
    })


@require_POST
@login_required
def vehicle_service_add(request, pk):
    """HR/MD: log a service record directly against a vehicle."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    vehicle = get_object_or_404(Vehicle, pk=pk)

    service_type = (request.POST.get('service_type') or '').strip()
    if not _valid_service_type(service_type):
        messages.error(request, 'Please choose a valid service type.')
        return redirect('vehicle_services', pk=pk)

    other_detail = (request.POST.get('other_detail') or '').strip()
    if service_type == 'other' and not other_detail:
        messages.error(request, 'Please specify the service in the "Other" field.')
        return redirect('vehicle_services', pk=pk)

    status = request.POST.get('status') or 'Completed'
    if status not in dict(VehicleService.STATUS_CHOICES):
        status = 'Completed'

    cost = (request.POST.get('cost') or '').strip()
    svc = VehicleService(
        vehicle=vehicle,
        service_type=service_type,
        other_detail=other_detail,
        status=status,
        service_date=request.POST.get('service_date') or None,
        notes=(request.POST.get('notes') or '').strip(),
        cost=cost or None,
        created_by=request.user,
    )
    if status in ('Approved', 'Completed'):
        svc.approved_by = request.user
        svc.approved_at = timezone.now()
    # HR logged this directly, so any cost entered is approved on the spot.
    if svc.cost is not None:
        svc.cost_approved = True
        svc.cost_approved_by = request.user
        svc.cost_approved_at = timezone.now()
    svc.save()
    messages.success(request, f"Service record added for '{vehicle.name}'.")
    return redirect('vehicle_services', pk=pk)


@require_POST
@login_required
def vehicle_service_process(request, pk):
    """HR/MD actions on a service request.

    Workflow: employee requests -> HR approves -> employee completes & records
    cost -> HR approves the cost.
      * approve       — Requested  -> Approved
      * reject        — Requested  -> Rejected
      * complete      — Approved   -> Completed (HR completing on the employee's
                        behalf; cost entered here is auto-approved)
      * approve_cost  — Completed  -> cost_approved = True (optionally editing cost)
    """
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    svc = get_object_or_404(VehicleService.objects.select_related('vehicle'), pk=pk)
    action = request.POST.get('action')

    if action == 'approve':
        svc.status = 'Approved'
        svc.approved_by = request.user
        svc.approved_at = timezone.now()
        svc.save()
        messages.success(request, 'Service request approved. The employee can now complete it and record the cost.')
    elif action == 'complete':
        svc.status = 'Completed'
        svc.service_date = request.POST.get('service_date') or svc.service_date or timezone.localdate()
        notes = (request.POST.get('notes') or '').strip()
        if notes:
            svc.notes = notes
        cost = (request.POST.get('cost') or '').strip()
        if cost:
            svc.cost = cost
            # HR entered the cost themselves, so it's approved on the spot.
            svc.cost_approved = True
            svc.cost_approved_by = request.user
            svc.cost_approved_at = timezone.now()
        if not svc.approved_by:
            svc.approved_by = request.user
            svc.approved_at = timezone.now()
        svc.save()
        messages.success(request, 'Service marked as completed and added to history.')
    elif action == 'approve_cost':
        if svc.status != 'Completed':
            messages.error(request, 'Only a completed service can have its cost approved.')
        else:
            cost = (request.POST.get('cost') or '').strip()
            if cost:
                svc.cost = cost
            if svc.cost is None:
                messages.error(request, 'Enter a cost before approving it.')
            else:
                svc.cost_approved = True
                svc.cost_approved_by = request.user
                svc.cost_approved_at = timezone.now()
                svc.save()
                messages.success(request, 'Cost approved.')
    elif action == 'reject':
        svc.status = 'Rejected'
        svc.approved_by = request.user
        svc.approved_at = timezone.now()
        svc.rejection_reason = (request.POST.get('rejection_reason') or '').strip()
        svc.save()
        messages.success(request, 'Service request rejected.')
    else:
        messages.error(request, 'Unknown action.')

    next_url = request.POST.get('next')
    if next_url == 'requests':
        return redirect('vehicle_service_requests')
    return redirect('vehicle_services', pk=svc.vehicle_id)


@require_POST
@login_required
def vehicle_service_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    svc = get_object_or_404(VehicleService, pk=pk)
    vehicle_id = svc.vehicle_id
    svc.delete()
    messages.success(request, 'Service record deleted.')
    return redirect('vehicle_services', pk=vehicle_id)


@login_required
def vehicle_service_requests(request):
    """HR/MD: queue of pending service requests across all vehicles."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    pending = VehicleService.objects.filter(status='Requested').select_related(
        'vehicle', 'requested_by'
    )
    # Completed services with a cost the employee entered, still awaiting HR sign-off.
    cost_pending = VehicleService.objects.filter(
        status='Completed', cost_approved=False, cost__isnull=False
    ).select_related('vehicle', 'requested_by')
    recent = VehicleService.objects.exclude(status='Requested').select_related(
        'vehicle', 'requested_by', 'approved_by'
    )[:30]
    return render(request, 'hr_de/vehicles/service_requests.html', {
        'pending':      pending,
        'cost_pending': cost_pending,
        'recent':       recent,
    })


@login_required
def my_vehicles(request):
    """Employee self-service: vehicles assigned to me + their service history."""
    employee = _employee_for_user(request.user)
    if employee is None:
        return render(request, 'hr_de/unauthorized.html', status=403)

    vehicles = Vehicle.objects.filter(employee=employee).prefetch_related('services')
    my_requests = VehicleService.objects.filter(
        requested_by=employee
    ).select_related('vehicle')

    # Consolidated service history across all of the employee's vehicles.
    service_history = VehicleService.objects.filter(
        vehicle__employee=employee
    ).select_related('vehicle').order_by('-service_date', '-created_at')

    # Only these types are offered to employees.
    requestable = [
        (v, l) for v, l in VehicleService.SERVICE_TYPE_CHOICES
        if v in VehicleService.EMPLOYEE_REQUESTABLE
    ]
    return render(request, 'hr_de/vehicles/my_vehicles.html', {
        'employee':        employee,
        'vehicles':        vehicles,
        'my_requests':     my_requests,
        'service_history': service_history,
        'requestable':     requestable,
        'today':           timezone.localdate(),
    })


@require_POST
@login_required
def request_vehicle_service(request, pk):
    """Employee submits a service request for one of their assigned vehicles."""
    employee = _employee_for_user(request.user)
    if employee is None:
        return render(request, 'hr_de/unauthorized.html', status=403)

    vehicle = get_object_or_404(Vehicle, pk=pk)
    if vehicle.employee_id != employee.id:
        messages.error(request, 'You can only request service for a vehicle assigned to you.')
        return redirect('my_vehicles')

    service_type = (request.POST.get('service_type') or '').strip()
    if service_type not in VehicleService.EMPLOYEE_REQUESTABLE:
        messages.error(request, 'Please choose a valid service type.')
        return redirect('my_vehicles')

    other_detail = (request.POST.get('other_detail') or '').strip()
    if service_type == 'other' and not other_detail:
        messages.error(request, 'Please specify the service in the "Other" field.')
        return redirect('my_vehicles')

    VehicleService.objects.create(
        vehicle=vehicle,
        service_type=service_type,
        other_detail=other_detail,
        status='Requested',
        notes=(request.POST.get('notes') or '').strip(),
        requested_by=employee,
    )
    messages.success(request, 'Your service request has been submitted to HR.')
    return redirect('my_vehicles')


@require_POST
@login_required
def complete_vehicle_service(request, pk):
    """Employee marks an HR-approved service as completed and records the cost.
    HR then approves the cost separately."""
    employee = _employee_for_user(request.user)
    if employee is None:
        return render(request, 'hr_de/unauthorized.html', status=403)

    svc = get_object_or_404(VehicleService.objects.select_related('vehicle'), pk=pk)
    # Must be the employee's own request or their assigned vehicle.
    if svc.requested_by_id != employee.id and svc.vehicle.employee_id != employee.id:
        messages.error(request, 'You can only update a service for your own vehicle.')
        return redirect('my_vehicles')

    if svc.status != 'Approved':
        messages.error(request, 'Only an HR-approved service can be marked completed.')
        return redirect('my_vehicles')

    cost = (request.POST.get('cost') or '').strip()
    if not cost:
        messages.error(request, 'Please enter the service cost.')
        return redirect('my_vehicles')

    svc.status       = 'Completed'
    svc.service_date = request.POST.get('service_date') or timezone.localdate()
    svc.cost         = cost
    extra_notes = (request.POST.get('notes') or '').strip()
    if extra_notes:
        svc.notes = extra_notes
    # Employee-entered cost still needs HR sign-off.
    svc.cost_approved = False
    svc.save()
    messages.success(request, 'Service marked completed. The cost is now pending HR approval.')
    return redirect('my_vehicles')


# ── Vehicle Bulk Upload (Excel) ──────────────────────────────────────────────

# (Excel header, note/instruction). NAME + CAR NUMBER are required; the rest
# are optional. OWNERSHIP defaults to "Company" when left blank.
_VEHICLE_BULK_COLUMNS = [
    ("NAME",             "Required · owner / vehicle name"),
    ("CAR NUMBER",       "Required · plate / registration no. (used to match on re-upload)"),
    ("MODEL",            "e.g. 2022"),
    ("CAR AND MODEL",    "e.g. Toyota Camry 2022"),
    ("COMPANY",          "Owning / registered company"),
    ("TRACKING",         "Provider / device / Yes"),
    ("TRACKING EXP DATE","YYYY-MM-DD"),
    ("STATE",            "e.g. Dubai / Abu Dhabi"),
    ("TRAFFIC CODE",     "Traffic file / code"),
    ("MORTGAGE",         "Bank name / None"),
    ("MULKIYA EXPIRY",   "YYYY-MM-DD"),
    ("OWNERSHIP",        "Company or Personal · blank = Company"),
    ("EMPLOYEE EMP ID",  "Optional · EMP ID of associated employee"),
]


@login_required
def vehicle_upload_template(request):
    """Return a pre-formatted .xlsx template for bulk-adding vehicles."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    note_fill   = PatternFill("solid", fgColor="EBF3FF")
    sample_fill = PatternFill("solid", fgColor="F0FFF4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    note_font   = Font(italic=True, color="1E3A5F", size=9)
    sample_font = Font(color="1E3A5F", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1 – headers
    for col_idx, (header, _) in enumerate(_VEHICLE_BULK_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, thin_border

    # Row 2 – notes
    for col_idx, (_, note) in enumerate(_VEHICLE_BULK_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.fill      = note_fill
        cell.font      = note_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border    = thin_border

    # Row 3 – sample
    sample = [
        "Ali Hassan", "A 12345", "2022", "Toyota Camry 2022", "ABC Trading LLC",
        "Yes", "2026-12-31", "Dubai", "TC-99887", "Emirates NBD",
        "2026-08-15", "Company", "EMP001",
    ]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill, cell.font, cell.border = sample_fill, sample_font, thin_border

    # Keep CAR NUMBER / TRAFFIC CODE / EMP ID as text so leading zeros survive
    text_cols = {"CAR NUMBER", "TRAFFIC CODE", "EMPLOYEE EMP ID"}
    for ci, (h, _) in enumerate(_VEHICLE_BULK_COLUMNS, start=1):
        if h in text_cols:
            for ri in range(3, 1001):
                ws.cell(row=ri, column=ci).number_format = "@"

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = "A4"
    col_widths = [20, 16, 10, 22, 22, 14, 16, 16, 14, 16, 16, 14, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="vehicle_upload_template.xlsx"'
    wb.save(response)
    return response


@login_required
def vehicle_bulk_upload(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == "POST" and request.FILES.get("excel_file"):
        import pandas as pd
        excel_file = request.FILES["excel_file"]
        try:
            df = pd.read_excel(
                excel_file, header=0, skiprows=[1],   # skip the notes row
                dtype={"CAR NUMBER": str, "TRAFFIC CODE": str, "EMPLOYEE EMP ID": str},
            )

            created_count = 0
            updated_count = 0
            error_rows    = []
            # map lower-cased input → canonical value, so "personal"/"PERSONAL"
            # all resolve to "Personal" (blank/unknown → Company).
            owner_lookup  = {v.lower(): v for v, _ in Vehicle.OWNERSHIP_CHOICES}

            def _str(row, key):
                val = row.get(key)
                if val is None or (isinstance(val, float) and pd.isnull(val)):
                    return ''
                s = str(val).strip()
                if s.lower() == 'nan':
                    return ''
                if s.endswith('.0') and s[:-2].isdigit():   # Excel numeric coercion
                    s = s[:-2]
                return s

            for index, row in df.iterrows():
                try:
                    name = _str(row, "NAME")
                    car_number = _str(row, "CAR NUMBER")
                    if not name and not car_number:
                        continue   # blank row
                    if not name or not car_number:
                        error_rows.append(f"Row {index + 3}: NAME and CAR NUMBER are both required — skipped.")
                        continue

                    ownership = owner_lookup.get(_str(row, "OWNERSHIP").lower(), "Company")

                    emp_id = _str(row, "EMPLOYEE EMP ID")
                    employee = None
                    if emp_id:
                        employee = Employee.objects.filter(emp_id__iexact=emp_id).first()
                        if employee is None:
                            error_rows.append(f"Row {index + 3}: no employee with EMP ID '{emp_id}' — association left blank.")

                    _, created = Vehicle.objects.update_or_create(
                        car_number=car_number,
                        defaults={
                            "name":              name,
                            "ownership":         ownership,
                            "model":             _str(row, "MODEL"),
                            "car_and_model":     _str(row, "CAR AND MODEL"),
                            "company":           _str(row, "COMPANY"),
                            "tracking":          _str(row, "TRACKING"),
                            "tracking_exp_date": _parse_date(row.get("TRACKING EXP DATE")),
                            "state":             _str(row, "STATE"),
                            "traffic_code":      _str(row, "TRAFFIC CODE"),
                            "mortgage":          _str(row, "MORTGAGE"),
                            "mulkiya_expiry":    _parse_date(row.get("MULKIYA EXPIRY")),
                            "employee":          employee,
                            "created_by":        request.user,
                        },
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    error_rows.append(f"Row {index + 3}: {e}")

            for err in error_rows:
                messages.error(request, err)
            if created_count or updated_count:
                messages.success(
                    request,
                    f"Upload complete — {created_count} added, {updated_count} updated."
                )
            elif not error_rows:
                messages.warning(request, "No vehicles found in the file.")
            return redirect("vehicle_upload")

        except Exception as e:
            messages.error(request, f"Upload failed: {e}")

    columns = [col for col, _ in _VEHICLE_BULK_COLUMNS]
    return render(request, "hr_de/vehicles/bulk_upload.html", {"columns": columns})


# ── Monthly Reviews (Head fills; HR/MD read) ─────────────────────────────────

_REVIEW_RATINGS = ['rating_task', 'rating_punctuality', 'rating_quality', 'rating_communication']


def _month_year_from(request, today):
    try:
        month = int(request.GET.get('month') or request.POST.get('month') or today.month)
    except (ValueError, TypeError):
        month = today.month
    try:
        year = int(request.GET.get('year') or request.POST.get('year') or today.year)
    except (ValueError, TypeError):
        year = today.year
    if not (1 <= month <= 12):
        month = today.month
    return month, year


def _review_week_from(request):
    try:
        week = int(request.GET.get('week') or request.POST.get('week') or 0)
    except (ValueError, TypeError):
        week = 0
    return week if week in (1, 2, 3, 4, 5) else 0


def _review_defaults_from_post(request):
    def _rating(key):
        try:
            v = int(request.POST.get(key) or 0)
        except (ValueError, TypeError):
            return None
        return v if 1 <= v <= 5 else None
    return {
        'rating_task':          _rating('rating_task'),
        'rating_punctuality':   _rating('rating_punctuality'),
        'rating_quality':       _rating('rating_quality'),
        'rating_communication': _rating('rating_communication'),
        'went_well':       (request.POST.get('went_well') or '').strip(),
        'concerns':        (request.POST.get('concerns') or '').strip(),
        'needs_attention': request.POST.get('needs_attention') == 'on',
        'reviewed_by':     request.user,
    }


@login_required
def review_team(request):
    """Department head sees their team + review status for a month (or a week
    of that month when ?week=1..5)."""
    dept = _head_department(request)
    if dept is None:
        return render(request, 'hr_de/unauthorized.html', status=403)

    import datetime as dt
    today = timezone.localdate()
    month, year = _month_year_from(request, today)
    week = _review_week_from(request)

    employees = list(
        Employee.objects.filter(department=dept, is_active=True)
        .exclude(user=request.user)          # heads don't review themselves
        .order_by('emp_name')
    )
    if week:
        reviews = {r.employee_id: r for r in WeeklyReview.objects.filter(
            employee__in=employees, month=month, year=year, week=week)}
    else:
        reviews = {r.employee_id: r for r in MonthlyReview.objects.filter(
            employee__in=employees, month=month, year=year)}
    rows = [{'employee': e, 'review': reviews.get(e.id)} for e in employees]
    done = sum(1 for r in rows if r['review'])

    months = [{'num': i, 'name': dt.date(2000, i, 1).strftime('%B')} for i in range(1, 13)]
    years  = list(range(today.year - 2, today.year + 1))

    return render(request, 'hr_de/reviews/team.html', {
        'department': dept,
        'rows':       rows,
        'month':      month,
        'year':       year,
        'week':       week,
        'weeks':      [1, 2, 3, 4, 5],
        'months':     months,
        'years':      years,
        'done':       done,
        'pending':    len(rows) - done,
    })


@login_required
def review_edit(request, emp_pk):
    """Head fills / updates a monthly OR weekly review (week via ?week=1..5)."""
    dept = _head_department(request)
    if dept is None:
        return render(request, 'hr_de/unauthorized.html', status=403)
    employee = get_object_or_404(Employee, pk=emp_pk, department=dept, is_active=True)

    import datetime as dt
    today = timezone.localdate()
    month, year = _month_year_from(request, today)
    week = _review_week_from(request)

    if week:
        review = WeeklyReview.objects.filter(employee=employee, month=month, year=year, week=week).first()
    else:
        review = MonthlyReview.objects.filter(employee=employee, month=month, year=year).first()

    if request.method == 'POST':
        defaults = _review_defaults_from_post(request)
        if week:
            WeeklyReview.objects.update_or_create(
                employee=employee, month=month, year=year, week=week, defaults=defaults)
        else:
            MonthlyReview.objects.update_or_create(
                employee=employee, month=month, year=year, defaults=defaults)
        messages.success(request, f"Review saved for {employee.emp_name}.")
        from django.urls import reverse
        url = f"{reverse('review_team')}?month={month}&year={year}"
        if week:
            url += f"&week={week}"
        return redirect(url)

    return render(request, 'hr_de/reviews/form.html', {
        'employee':     employee,
        'review':       review,
        'month':        month,
        'year':         year,
        'week':         week,
        'month_name':   dt.date(2000, month, 1).strftime('%B'),
        'period_word':  'week' if week else 'month',
        'rating_range': [1, 2, 3, 4, 5],
    })


@login_required
def review_employee(request, emp_pk):
    """HR / MD: one employee's monthly + weekly reviews for a chosen month."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    employee = get_object_or_404(Employee, pk=emp_pk)

    import datetime as dt
    today = timezone.localdate()
    month, year = _month_year_from(request, today)

    monthly = MonthlyReview.objects.filter(
        employee=employee, month=month, year=year).select_related('reviewed_by').first()
    weeklies = list(
        WeeklyReview.objects.filter(employee=employee, month=month, year=year)
        .select_related('reviewed_by').order_by('week')
    )

    months = [{'num': i, 'name': dt.date(2000, i, 1).strftime('%B')} for i in range(1, 13)]
    years  = list(range(today.year - 2, today.year + 1))

    return render(request, 'hr_de/reviews/employee.html', {
        'employee':   employee,
        'month':      month,
        'year':       year,
        'month_name': dt.date(2000, month, 1).strftime('%B'),
        'months':     months,
        'years':      years,
        'monthly':    monthly,
        'weeklies':   weeklies,
    })


@login_required
def review_overview(request):
    """HR / MD read all reviews for a month — with department & flag filters."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    import datetime as dt
    today = timezone.localdate()
    month, year = _month_year_from(request, today)

    reviews = (MonthlyReview.objects
               .filter(month=month, year=year)
               .select_related('employee', 'employee__department', 'reviewed_by'))

    dept_filter = (request.GET.get('department') or '').strip()
    if dept_filter:
        reviews = reviews.filter(employee__department_id=dept_filter)
    flagged_only = request.GET.get('flagged') == '1'
    if flagged_only:
        reviews = reviews.filter(needs_attention=True)

    reviews = list(reviews)
    flagged_count = sum(1 for r in reviews if r.needs_attention)

    months = [{'num': i, 'name': dt.date(2000, i, 1).strftime('%B')} for i in range(1, 13)]
    years  = list(range(today.year - 2, today.year + 1))

    return render(request, 'hr_de/reviews/overview.html', {
        'reviews':       reviews,
        'month':         month,
        'year':          year,
        'months':        months,
        'years':         years,
        'departments':   Department.objects.all().order_by('name'),
        'selected_dept': dept_filter,
        'flagged_only':  flagged_only,
        'flagged_count': flagged_count,
        'total':         len(reviews),
    })


# ── Management Details (owners/directors + families — HR & MD only) ──────────

_MGMT_TEXT_FIELDS = [
    'name', 'designation', 'nationality', 'phone', 'email',
    'eid_number', 'passport_number', 'dl_number', 'visa_number',
]
_MGMT_DATE_FIELDS = [
    'dob', 'eid_expiry', 'passport_expiry', 'dl_expiry', 'visa_expiry',
]
_MGMT_FILE_FIELDS = [
    'photo', 'eid_document', 'passport_document', 'dl_document', 'visa_document',
]


def _apply_management_post(member, request):
    """Copy submitted fields onto a ManagementMember (shared by add & edit)."""
    for f in _MGMT_TEXT_FIELDS:
        setattr(member, f, (request.POST.get(f) or '').strip())
    for f in _MGMT_DATE_FIELDS:
        setattr(member, f, request.POST.get(f) or None)

    visa_type = (request.POST.get('visa_type') or '').strip()
    member.visa_type = visa_type if visa_type in dict(ManagementMember.VISA_TYPE_CHOICES) else ''

    for f in _MGMT_FILE_FIELDS:
        if request.FILES.get(f):
            setattr(member, f, request.FILES[f])


@login_required
def management_list(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    members = ManagementMember.objects.filter(head__isnull=True).prefetch_related('family')

    query = (request.GET.get('q') or '').strip()
    if query:
        members = members.filter(
            Q(name__icontains=query) |
            Q(designation__icontains=query) |
            Q(nationality__icontains=query) |
            Q(family__name__icontains=query)
        ).distinct()

    return render(request, 'hr_de/management/list.html', {
        'members': members,
        'query':   query,
    })


@login_required
def management_detail(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    member = get_object_or_404(
        ManagementMember.objects.select_related('head').prefetch_related('family', 'travels', 'country_visas'), pk=pk
    )
    return render(request, 'hr_de/management/detail.html', {
        'member':             member,
        'today':              timezone.localdate(),
        'country_visa_types': CountryVisa.VISA_TYPE_CHOICES,
    })


@login_required
def management_add(request):
    """Add a new management person (head)."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Full name is required.')
            return redirect('management_add')
        member = ManagementMember(relation='self', created_by=request.user)
        _apply_management_post(member, request)
        member.relation = 'self'
        member.save()
        messages.success(request, f"Management person '{member.name}' added.")
        return redirect('management_detail', pk=member.pk)

    return render(request, 'hr_de/management/form.html', {
        'visa_choices':     ManagementMember.VISA_TYPE_CHOICES,
        'relation_choices': ManagementMember.RELATION_CHOICES,
        'is_family':        False,
    })


@login_required
def management_family_add(request, head_pk):
    """Add a family member under a management person."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    head = get_object_or_404(ManagementMember, pk=head_pk, head__isnull=True)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Full name is required.')
            return redirect('management_family_add', head_pk=head.pk)
        member = ManagementMember(head=head, created_by=request.user)
        _apply_management_post(member, request)
        relation = (request.POST.get('relation') or '').strip()
        member.relation = relation if relation in dict(ManagementMember.RELATION_CHOICES) and relation != 'self' else 'other'
        member.save()
        messages.success(request, f"Family member '{member.name}' added under {head.name}.")
        return redirect('management_detail', pk=head.pk)

    return render(request, 'hr_de/management/form.html', {
        'visa_choices':     ManagementMember.VISA_TYPE_CHOICES,
        'relation_choices': [c for c in ManagementMember.RELATION_CHOICES if c[0] != 'self'],
        'is_family':        True,
        'head':             head,
    })


@login_required
def management_edit(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    member = get_object_or_404(ManagementMember, pk=pk)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Full name is required.')
            return redirect('management_edit', pk=pk)
        _apply_management_post(member, request)
        if member.is_family:
            relation = (request.POST.get('relation') or '').strip()
            member.relation = relation if relation in dict(ManagementMember.RELATION_CHOICES) and relation != 'self' else 'other'
        else:
            member.relation = 'self'
        member.save()
        messages.success(request, f"'{member.name}' updated.")
        return redirect('management_detail', pk=member.pk)

    relation_choices = ManagementMember.RELATION_CHOICES
    if member.is_family:
        relation_choices = [c for c in ManagementMember.RELATION_CHOICES if c[0] != 'self']
    return render(request, 'hr_de/management/form.html', {
        'member':           member,
        'visa_choices':     ManagementMember.VISA_TYPE_CHOICES,
        'relation_choices': relation_choices,
        'is_family':        member.is_family,
        'head':             member.head,
    })


@require_POST
@login_required
def management_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    member = get_object_or_404(ManagementMember, pk=pk)
    name = member.name
    head_pk = member.head_id
    member.delete()
    messages.success(request, f"'{name}' deleted.")
    if head_pk:
        return redirect('management_detail', pk=head_pk)
    return redirect('management_list')


@require_POST
@login_required
def travel_add(request, member_pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    member = get_object_or_404(ManagementMember, pk=member_pk)
    destination = (request.POST.get('destination') or '').strip()
    if not destination:
        messages.error(request, 'Destination is required for a travel record.')
        return redirect('management_detail', pk=member.pk)
    TravelRecord.objects.create(
        member=member,
        destination=destination,
        purpose=(request.POST.get('purpose') or '').strip(),
        departure_date=request.POST.get('departure_date') or None,
        return_date=request.POST.get('return_date') or None,
        notes=(request.POST.get('notes') or '').strip(),
        document=request.FILES.get('document'),
        created_by=request.user,
    )
    messages.success(request, 'Travel record added.')
    return redirect('management_detail', pk=member.pk)


@require_POST
@login_required
def travel_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    travel = get_object_or_404(TravelRecord, pk=pk)
    member_pk = travel.member_id
    travel.delete()
    messages.success(request, 'Travel record deleted.')
    return redirect('management_detail', pk=member_pk)


@require_POST
@login_required
def country_visa_add(request, member_pk):
    """Add a visa the member holds for another country."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    member = get_object_or_404(ManagementMember, pk=member_pk)
    country = (request.POST.get('country') or '').strip()
    if not country:
        messages.error(request, 'Country is required for a country visa.')
        return redirect('management_detail', pk=member.pk)
    visa_type = (request.POST.get('visa_type') or '').strip()
    if visa_type not in dict(CountryVisa.VISA_TYPE_CHOICES):
        visa_type = ''
    CountryVisa.objects.create(
        member=member,
        country=country,
        visa_type=visa_type,
        number=(request.POST.get('number') or '').strip(),
        issue_date=request.POST.get('issue_date') or None,
        expiry=request.POST.get('expiry') or None,
        document=request.FILES.get('document'),
        created_by=request.user,
    )
    messages.success(request, f'{country} visa added.')
    return redirect('management_detail', pk=member.pk)


@require_POST
@login_required
def country_visa_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    visa = get_object_or_404(CountryVisa, pk=pk)
    member_pk = visa.member_id
    visa.delete()
    messages.success(request, 'Country visa deleted.')
    return redirect('management_detail', pk=member_pk)


# ── Company Properties (assets assignable to employees — HR & MD only) ────────

def _apply_property_post(prop, request):
    """Copy submitted fields onto a CompanyProperty (shared by add & edit)."""
    prop.name          = (request.POST.get('name') or '').strip()
    category           = (request.POST.get('category') or '').strip()
    prop.category      = category if category in dict(CompanyProperty.CATEGORY_CHOICES) else ''
    prop.serial_number = (request.POST.get('serial_number') or '').strip()
    prop.description   = (request.POST.get('description') or '').strip()
    prop.purchase_date = request.POST.get('purchase_date') or None
    prop.value         = (request.POST.get('value') or '').strip() or None

    emp_id = request.POST.get('assigned_to')
    prop.assigned_to = Employee.objects.filter(pk=emp_id).first() if emp_id else None
    # Assignment date only applies when assigned to someone.
    if prop.assigned_to:
        prop.assigned_on = request.POST.get('assigned_on') or prop.assigned_on or timezone.localdate()
    else:
        prop.assigned_on = None

    if request.FILES.get('document'):
        prop.document = request.FILES['document']


@login_required
def property_list(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    properties = CompanyProperty.objects.select_related('assigned_to').all()

    query = (request.GET.get('q') or '').strip()
    if query:
        properties = properties.filter(
            Q(name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(description__icontains=query) |
            Q(assigned_to__emp_name__icontains=query)
        ).distinct()

    category = (request.GET.get('category') or '').strip()
    if category in dict(CompanyProperty.CATEGORY_CHOICES):
        properties = properties.filter(category=category)

    assigned = (request.GET.get('assigned') or '').strip()
    if assigned == 'assigned':
        properties = properties.filter(assigned_to__isnull=False)
    elif assigned == 'available':
        properties = properties.filter(assigned_to__isnull=True)

    return render(request, 'hr_de/properties/list.html', {
        'properties':        properties,
        'query':             query,
        'category':          category,
        'assigned':          assigned,
        'category_choices':  CompanyProperty.CATEGORY_CHOICES,
        'assigned_count':    CompanyProperty.objects.filter(assigned_to__isnull=False).count(),
        'available_count':   CompanyProperty.objects.filter(assigned_to__isnull=True).count(),
    })


@login_required
def property_add(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Property name is required.')
            return redirect('property_add')
        prop = CompanyProperty(created_by=request.user)
        _apply_property_post(prop, request)
        prop.save()
        messages.success(request, f"Property '{prop.name}' added.")
        return redirect('property_list')

    return render(request, 'hr_de/properties/form.html', {
        'category_choices': CompanyProperty.CATEGORY_CHOICES,
        'employees':        Employee.objects.filter(is_active=True).order_by('emp_name'),
    })


@login_required
def property_edit(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    prop = get_object_or_404(CompanyProperty, pk=pk)

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Property name is required.')
            return redirect('property_edit', pk=pk)
        _apply_property_post(prop, request)
        prop.save()
        messages.success(request, f"Property '{prop.name}' updated.")
        return redirect('property_list')

    return render(request, 'hr_de/properties/form.html', {
        'property':         prop,
        'category_choices': CompanyProperty.CATEGORY_CHOICES,
        'employees':        Employee.objects.filter(is_active=True).order_by('emp_name'),
    })


@require_POST
@login_required
def property_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    prop = get_object_or_404(CompanyProperty, pk=pk)
    name = prop.name
    prop.delete()
    messages.success(request, f"Property '{name}' deleted.")
    return redirect('property_list')


# ── Memos (official memorandums on company letterhead — HR & MD only) ─────────

def _next_memo_ref():
    """Generate the next reference like JN/ADMIN/202603 (year + running number)."""
    year = timezone.now().year
    seq = Memo.objects.filter(memo_date__year=year).count() + 1
    return f"JN/ADMIN/{year}{seq:02d}"


def _resolve_memo_recipient(request):
    """Determine to_text / employee / department from the submitted memo type."""
    memo_type  = request.POST.get('memo_type') or 'general'
    employee   = None
    department = None
    if memo_type == 'warning_employee':
        employee = Employee.objects.filter(pk=request.POST.get('employee')).first()
        to_text  = employee.emp_name if employee else (request.POST.get('to_text') or '')
    elif memo_type == 'warning_department':
        department = Department.objects.filter(pk=request.POST.get('department')).first()
        to_text    = department.name if department else (request.POST.get('to_text') or '')
    else:
        to_text = (request.POST.get('to_text') or 'All Staff').strip() or 'All Staff'
    return memo_type, to_text, employee, department


@login_required
def memo_list(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    memos = Memo.objects.select_related('employee', 'department', 'created_by').all()
    query = (request.GET.get('q') or '').strip()
    if query:
        memos = memos.filter(
            Q(ref_no__icontains=query) | Q(subject__icontains=query) |
            Q(to_text__icontains=query) | Q(body__icontains=query)
        )
    mtype = (request.GET.get('memo_type') or '').strip()
    if mtype in dict(Memo.MEMO_TYPE_CHOICES):
        memos = memos.filter(memo_type=mtype)
    return render(request, 'hr_de/memos/list.html', {
        'memos': memos,
        'query': query,
        'memo_type': mtype,
        'type_choices': Memo.MEMO_TYPE_CHOICES,
    })


@login_required
def memo_add(request):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)

    if request.method == 'POST':
        subject = (request.POST.get('subject') or '').strip()
        body    = (request.POST.get('body') or '').strip()
        if not subject or not body:
            messages.error(request, 'Subject and body are required.')
            return redirect('memo_add')
        memo_type, to_text, employee, department = _resolve_memo_recipient(request)
        memo = Memo(
            memo_type=memo_type,
            ref_no=(request.POST.get('ref_no') or '').strip() or _next_memo_ref(),
            to_text=to_text,
            employee=employee,
            department=department,
            memo_date=request.POST.get('memo_date') or timezone.localdate(),
            subject=subject,
            body=body,
            signatory=(request.POST.get('signatory') or 'HR/ADMIN DIVISION').strip(),
            signature=request.FILES.get('signature'),
            created_by=request.user,
        )
        memo.save()
        messages.success(request, f"Memo '{memo.ref_no}' created.")
        return redirect('memo_list')

    return render(request, 'hr_de/memos/form.html', {
        'type_choices': Memo.MEMO_TYPE_CHOICES,
        'employees':    Employee.objects.filter(is_active=True).order_by('emp_name'),
        'departments':  Department.objects.all().order_by('name'),
        'suggested_ref': _next_memo_ref(),
        'today':        timezone.localdate().isoformat(),
    })


@login_required
def memo_edit(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    memo = get_object_or_404(Memo, pk=pk)

    if request.method == 'POST':
        subject = (request.POST.get('subject') or '').strip()
        body    = (request.POST.get('body') or '').strip()
        if not subject or not body:
            messages.error(request, 'Subject and body are required.')
            return redirect('memo_edit', pk=pk)
        memo_type, to_text, employee, department = _resolve_memo_recipient(request)
        memo.memo_type  = memo_type
        memo.ref_no     = (request.POST.get('ref_no') or '').strip() or memo.ref_no
        memo.to_text    = to_text
        memo.employee   = employee
        memo.department = department
        memo.memo_date  = request.POST.get('memo_date') or memo.memo_date
        memo.subject    = subject
        memo.body       = body
        memo.signatory  = (request.POST.get('signatory') or 'HR/ADMIN DIVISION').strip()
        if request.FILES.get('signature'):
            memo.signature = request.FILES['signature']
        memo.save()
        messages.success(request, f"Memo '{memo.ref_no}' updated.")
        return redirect('memo_list')

    return render(request, 'hr_de/memos/form.html', {
        'memo':         memo,
        'type_choices': Memo.MEMO_TYPE_CHOICES,
        'employees':    Employee.objects.filter(is_active=True).order_by('emp_name'),
        'departments':  Department.objects.all().order_by('name'),
        'today':        timezone.localdate().isoformat(),
    })


@require_POST
@login_required
def memo_delete(request, pk):
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    memo = get_object_or_404(Memo, pk=pk)
    ref = memo.ref_no
    memo.delete()
    messages.success(request, f"Memo '{ref}' deleted.")
    return redirect('memo_list')


@login_required
def memo_pdf(request, pk):
    """Render the memo as a PDF on the company letterhead (exact template design)."""
    if not _hr_or_md(request):
        return render(request, 'hr_de/unauthorized.html', status=403)
    memo = get_object_or_404(Memo, pk=pk)

    import os, html as _html
    from io import BytesIO
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Image as RLImage, Frame

    PW, PH = A4                       # 595.276 x 841.89
    IMG_W, IMG_H = 1241, 1755         # letterhead template pixel size
    sx, sy = PW / IMG_W, PH / IMG_H

    def X(px):  # pixel-x → points
        return px * sx

    def Y(py):  # pixel-y (from top) → points (baseline from bottom)
        return PH - py * sy

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    # Letterhead background (logo + labels + footer already printed on it).
    letterhead = os.path.join(settings.MEDIA_ROOT, 'HR MEMO SAMPLE.JPG')
    if os.path.exists(letterhead):
        c.drawImage(letterhead, 0, 0, width=PW, height=PH,
                    preserveAspectRatio=False, mask='auto')

    # Header values, aligned to the pre-printed Ref/To/Date/Re labels.
    c.setFont('Helvetica', 13)
    val_x = X(300)
    c.drawString(val_x, Y(328), memo.ref_no or '')
    c.drawString(val_x, Y(400), memo.to_text or '')
    c.drawString(val_x, Y(470), memo.memo_date.strftime('%d %B %Y') if memo.memo_date else '')

    # 'Re' subject may be long — wrap it inside a frame beside the Re label.
    # Frame content anchors at the TOP, so top edge is set just above the label.
    re_style = ParagraphStyle('re', fontName='Helvetica', fontSize=13, leading=17)
    re_para  = Paragraph(_html.escape(memo.subject or ''), re_style)
    re_top_px, re_bot_px = 515, 625
    re_frame = Frame(val_x, Y(re_bot_px), PW - val_x - X(90), Y(re_top_px) - Y(re_bot_px),
                     leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, showBoundary=0)
    re_frame.addFromList([re_para], c)

    # Body + closing, flowed in a frame beneath the header block.
    body_style = ParagraphStyle('body', fontName='Helvetica', fontSize=12.5, leading=21)
    close_style = ParagraphStyle('close', fontName='Helvetica', fontSize=12.5, leading=18)
    sign_style = ParagraphStyle('sign', fontName='Helvetica-Oblique', fontSize=12.5, leading=16)

    body_html = _html.escape(memo.body or '').replace('\n', '<br/>')
    story = [
        Paragraph(body_html, body_style),
        Spacer(1, 34),
        Paragraph('Thanks and regards,', close_style),
    ]
    if memo.signature:
        try:
            sig_path = memo.signature.path
            if os.path.exists(sig_path):
                img = RLImage(sig_path)
                # scale to max ~130pt wide / 60pt tall, keep aspect
                iw, ih = img.imageWidth, img.imageHeight
                scale = min(130.0 / iw, 60.0 / ih)
                img.drawWidth, img.drawHeight = iw * scale, ih * scale
                img.hAlign = 'LEFT'   # default is CENTER
                story += [Spacer(1, 4), img]
        except Exception:
            pass
    story += [Spacer(1, 6), Paragraph(memo.signatory or 'HR/ADMIN DIVISION', sign_style)]

    body_top = Y(640)
    frame = Frame(X(150), Y(1560), PW - X(150) - X(90), body_top - Y(1560),
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, showBoundary=0)
    frame.addFromList(story, c)

    c.showPage()
    c.save()
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    fname = (memo.ref_no or 'memo').replace('/', '-')
    resp['Content-Disposition'] = f'inline; filename="{fname}.pdf"'
    return resp
